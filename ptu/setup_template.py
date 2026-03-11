"""
Generador de Plantilla: Calculadora de PTU
==========================================
Crea un archivo Excel (.xlsx) con 8 hojas:
  - Instrucciones: Guia de uso paso a paso
  - Config: Parametros del ejercicio (UMA, SMG, topes)
  - Tarifa: Tarifa Art. 96 LISR mensual
  - Datos: Hoja principal (empresa + tabla de empleados + calculo PTU)
  - Calculo_ISR: ISR Art. 96 y Art. 174 con comparacion
  - Resumen: Totales generales y por empleado
  - Recibos: Template de recibo para impresion
  - Pre_Nomina: Borrador CFDI para timbrado

Marca: Columna 13
Uso:
    python3 ptu/setup_template.py
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.page import PrintPageSetup
from pathlib import Path


def add_named_range(wb, name, attr_text):
    """Add a named range compatible with openpyxl 3.x."""
    dn = DefinedName(name, attr_text=attr_text)
    wb.defined_names.add(dn)


# ==============================================================================
# CONSTANTES DE ESTILO
# ==============================================================================
FONT_FAMILY = "Aptos"
COLOR_HEADER_BG = "2C3E50"       # Azul oscuro (navy)
COLOR_HEADER_FG = "FFFFFF"       # Blanco
COLOR_ACCENT = "3498DB"          # Azul
COLOR_LIGHT_BG = "F5F7F9"       # Gris muy claro (filas alternas)
COLOR_SECTION_BG = "EBF5FB"     # Azul muy claro
COLOR_WARNING = "E74C3C"        # Rojo
COLOR_SUCCESS = "1E6B3A"        # Verde oscuro (columnas calculadas)
COLOR_CALC_HEADER = "27AE60"    # Verde (header columnas calculadas)
COLOR_MUTED = "7F8C8D"          # Gris

FONT_HEADER = Font(name=FONT_FAMILY, bold=True, color=COLOR_HEADER_FG, size=11)
FONT_CALC_HEADER = Font(name=FONT_FAMILY, bold=True, color=COLOR_HEADER_FG, size=11)
FONT_SECTION = Font(name=FONT_FAMILY, bold=True, color=COLOR_HEADER_BG, size=13)
FONT_SUBSECTION = Font(name=FONT_FAMILY, bold=True, color=COLOR_ACCENT, size=11)
FONT_NORMAL = Font(name=FONT_FAMILY, size=11)
FONT_SMALL = Font(name=FONT_FAMILY, size=9, color=COLOR_MUTED)
FONT_TITLE = Font(name=FONT_FAMILY, bold=True, size=18, color=COLOR_HEADER_BG)
FONT_SUBTITLE = Font(name=FONT_FAMILY, size=12, color=COLOR_MUTED)
FONT_LABEL = Font(name=FONT_FAMILY, bold=True, size=11, color=COLOR_HEADER_BG)
FONT_VALUE = Font(name=FONT_FAMILY, size=11)
FONT_FORMULA_LABEL = Font(name=FONT_FAMILY, bold=True, size=11, color=COLOR_SUCCESS)

FILL_HEADER = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
FILL_CALC_HEADER = PatternFill(start_color=COLOR_CALC_HEADER, end_color=COLOR_CALC_HEADER, fill_type="solid")
FILL_LIGHT = PatternFill(start_color=COLOR_LIGHT_BG, end_color=COLOR_LIGHT_BG, fill_type="solid")
FILL_SECTION = PatternFill(start_color=COLOR_SECTION_BG, end_color=COLOR_SECTION_BG, fill_type="solid")
FILL_SUCCESS_LIGHT = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")

THIN_BORDER = Border(bottom=Side(style="thin", color="BDC3C7"))
DOUBLE_TOP = Border(top=Side(style="double", color=COLOR_HEADER_BG))

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# Layout constants
NUM_EMPLOYEES = 50
DAT_HEADER_ROW = 13
DAT_DATA_START = 14
DAT_DATA_END = DAT_DATA_START + NUM_EMPLOYEES - 1  # 63
DAT_TOTALS_ROW = DAT_DATA_END + 1                  # 64

ISR_HEADER_ROW = 1
ISR_DATA_START = 2
ISR_DATA_END = ISR_DATA_START + NUM_EMPLOYEES - 1   # 51

OUTPUT_FILE = Path(__file__).parent / "CalculadoraPTU_LFT.xlsx"


# ==============================================================================
# DATOS CONSTANTES
# ==============================================================================

# Tarifa Art. 96 LISR Mensual 2024-2025
TARIFA_2024_2025 = [
    (0.01,       746.04,      0.00,       0.0192),
    (746.05,     6332.05,     14.32,      0.0640),
    (6332.06,    11128.01,    371.83,     0.1088),
    (11128.02,   12935.82,    893.63,     0.1600),
    (12935.83,   15487.71,    1182.88,    0.1792),
    (15487.72,   31236.49,    1640.18,    0.2136),
    (31236.50,   49233.00,    5004.12,    0.2352),
    (49233.01,   93993.90,    9236.89,    0.3000),
    (93993.91,   125325.20,   22665.17,   0.3200),
    (125325.21,  375975.61,   32691.18,   0.3400),
    (375975.62,  9999999999,  117912.32,  0.3500),
]

# Tarifa Art. 96 LISR Mensual 2026 (actualizada factor ~1.1321)
TARIFA_2026 = [
    (0.01,       844.59,      0.00,       0.0192),
    (844.60,     7168.51,     16.22,      0.0640),
    (7168.52,    12598.02,    420.95,     0.1088),
    (12598.03,   14644.64,    1011.68,    0.1600),
    (14644.65,   17533.64,    1339.14,    0.1792),
    (17533.65,   35362.83,    1856.84,    0.2136),
    (35362.84,   55736.68,    5665.16,    0.2352),
    (55736.69,   106410.50,   10457.09,   0.3000),
    (106410.51,  141880.66,   25659.23,   0.3200),
    (141880.67,  425641.99,   37009.69,   0.3400),
    (425642.00,  9999999999,  133488.54,  0.3500),
]

# Tabla UMA / SMG por año: (UMA_diaria, SMG_diaria, SMG_frontera)
UMA_SMG_DATA = [
    (2016, 73.04,  73.04,  73.04),
    (2017, 75.49,  80.04,  80.04),
    (2018, 80.60,  88.36,  88.36),
    (2019, 84.49,  102.68, 176.72),
    (2020, 86.88,  123.22, 185.56),
    (2021, 89.62,  141.70, 213.39),
    (2022, 96.22,  172.87, 260.34),
    (2023, 103.74, 207.44, 312.41),
    (2024, 108.57, 248.93, 374.89),
    (2025, 113.14, 278.80, 419.88),
    (2026, 117.31, 315.04, 440.87),
]


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def apply_header_style(ws, row, col_start, col_end, fill=None):
    """Apply header styling to a row range."""
    use_fill = fill or FILL_HEADER
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER if fill is None else FONT_CALC_HEADER
        cell.fill = use_fill
        cell.alignment = ALIGN_CENTER


def set_col_widths(ws, widths_dict):
    """Set column widths from a dict {col_letter: width}."""
    for col_letter, w in widths_dict.items():
        ws.column_dimensions[col_letter].width = w


# ==============================================================================
# INSTRUCCIONES SHEET
# ==============================================================================

_INS_DARK = "2C3E50"
_INS_BLUE = "3498DB"
_INS_GRAY = "5D6D7E"
_INS_GREEN = "1E6B3A"
_INS_TABLE_BG = "F4F6F8"


def _ins_write_header(ws, r, text):
    cell = ws.cell(row=r, column=2, value=text)
    cell.font = Font(name=FONT_FAMILY, size=15, bold=True, color=_INS_DARK)
    return r + 1


def _ins_write_step(ws, r, numero, titulo):
    cell = ws.cell(row=r, column=2, value=f"Paso {numero}: {titulo}")
    cell.font = Font(name=FONT_FAMILY, size=13, bold=True, color=_INS_BLUE)
    return r + 1


def _ins_write_text(ws, r, text):
    cell = ws.cell(row=r, column=2, value=text)
    cell.font = Font(name=FONT_FAMILY, size=11, color=_INS_GRAY)
    return r + 1


def _ins_write_bullet(ws, r, label, description):
    full = f"\u2022  {label} \u2014 {description}"
    cell = ws.cell(row=r, column=2, value=full)
    cell.font = Font(name=FONT_FAMILY, size=11, color=_INS_GRAY)
    return r + 1


def _ins_write_note(ws, r, text):
    cell = ws.cell(row=r, column=2, value=f"\u26A0  {text}")
    cell.font = Font(name=FONT_FAMILY, size=10, italic=True, color=_INS_GREEN)
    ws.cell(row=r, column=2).fill = PatternFill(
        start_color=_INS_TABLE_BG, end_color=_INS_TABLE_BG, fill_type="solid"
    )
    return r + 1


def _ins_write_table_row(ws, r, col1, col2):
    cell = ws.cell(row=r, column=2, value=f"{col1}    \u2192    {col2}")
    cell.font = Font(name=FONT_FAMILY, size=11, color=_INS_GRAY)
    if r % 2 == 0:
        cell.fill = PatternFill(
            start_color=_INS_TABLE_BG, end_color=_INS_TABLE_BG, fill_type="solid"
        )
    return r + 1


def _ins_write_separator(ws, r):
    cell = ws.cell(row=r, column=2, value="\u2500" * 90)
    cell.font = Font(name=FONT_FAMILY, size=8, color="BDC3C7")
    return r + 1


def create_instrucciones_sheet(wb):
    ws = wb.create_sheet("Instrucciones")
    ws.sheet_properties.tabColor = COLOR_ACCENT
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 120

    r = 2
    # Brand
    cell = ws.cell(row=r, column=2, value="Columna 13")
    cell.font = Font(name=FONT_FAMILY, size=11, color=_INS_BLUE)
    r += 1

    # Title
    cell = ws.cell(row=r, column=2, value="Calculadora de PTU")
    cell.font = Font(name=FONT_FAMILY, size=22, bold=True, color=_INS_DARK)
    r += 1

    cell = ws.cell(row=r, column=2,
                   value="Participaci\u00f3n de los Trabajadores en las Utilidades")
    cell.font = Font(name=FONT_FAMILY, size=14, color=_INS_GRAY)
    r += 1

    cell = ws.cell(row=r, column=2,
                   value="Art\u00edculos 117 al 131 de la Ley Federal del Trabajo | Versi\u00f3n 1.0")
    cell.font = Font(name=FONT_FAMILY, size=10, italic=True, color=_INS_GRAY)
    r += 2

    r = _ins_write_separator(ws, r)
    r += 1

    # Description
    r = _ins_write_header(ws, r, "Descripci\u00f3n")
    r = _ins_write_text(ws, r,
        "Esta plantilla calcula la Participaci\u00f3n de los Trabajadores en las Utilidades (PTU) "
        "conforme a la Ley Federal del Trabajo. Distribuye la PTU en dos partes iguales "
        "(50% d\u00edas trabajados, 50% salarios devengados), aplica el tope del Art. 127 fr. VIII "
        "y calcula el ISR a retener por dos m\u00e9todos (Art. 96 LISR y Art. 174 RLISR).")
    r += 1

    # Sheets
    r = _ins_write_header(ws, r, "Hojas del libro")
    sheets = [
        ("Instrucciones", "Esta gu\u00eda de uso"),
        ("Config", "Ejercicio fiscal, UMA, SMG, par\u00e1metros"),
        ("Tarifa", "Tarifa Art. 96 LISR mensual (2024-2026)"),
        ("Datos", "Datos de empresa y tabla de empleados con c\u00e1lculo de PTU"),
        ("C\u00e1lculo_ISR", "ISR por Art. 96 y Art. 174, con comparaci\u00f3n autom\u00e1tica"),
        ("Resumen", "Totales generales y por empleado"),
        ("Recibos", "Plantilla de recibo para impresi\u00f3n / firma"),
        ("Pre_N\u00f3mina", "Borrador CFDI para timbrado de n\u00f3mina extraordinaria"),
    ]
    for name, desc in sheets:
        r = _ins_write_bullet(ws, r, name, desc)
    r += 1

    r = _ins_write_separator(ws, r)
    r += 1

    # Steps
    r = _ins_write_header(ws, r, "Flujo de trabajo")
    r += 1

    r = _ins_write_step(ws, r, 1, "Configurar par\u00e1metros (hoja Config)")
    r = _ins_write_text(ws, r,
        "Verifica el Ejercicio fiscal (default 2025). La UMA y SMG se autocompletan.")
    r = _ins_write_text(ws, r,
        "Selecciona si usas UMA o SMG para la exenci\u00f3n de ISR (criterio SAT vs PRODECON).")
    r += 1

    r = _ins_write_step(ws, r, 2, "Capturar datos de la empresa (hoja Datos)")
    r = _ins_write_text(ws, r,
        "Llena: Nombre de la empresa, RFC, Utilidad Fiscal y PTU No Cobrada de ejercicios anteriores.")
    r = _ins_write_text(ws, r,
        "La PTU Generada (10%), PTU a Repartir y la divisi\u00f3n 50/50 se calculan autom\u00e1ticamente.")
    r += 1

    r = _ins_write_step(ws, r, 3, "Registrar empleados (hoja Datos, fila 14+)")
    r = _ins_write_text(ws, r,
        "Captura los datos de cada empleado en las columnas A-O (fondo azul = entrada manual).")
    r = _ins_write_text(ws, r,
        "Las columnas P-X (fondo verde) se calculan solas: factor, PTU por d\u00edas/salarios, tope y PTU real.")
    r += 1

    r = _ins_write_step(ws, r, 4, "Revisar el c\u00e1lculo de ISR (hoja C\u00e1lculo_ISR)")
    r = _ins_write_text(ws, r,
        "La hoja calcula autom\u00e1ticamente el ISR por ambos m\u00e9todos:")
    r = _ins_write_bullet(ws, r, "Art. 96 LISR",
        "M\u00e9todo de Ley. Retenci\u00f3n alta, sin diferencias en ajuste anual.")
    r = _ins_write_bullet(ws, r, "Art. 174 RLISR",
        "M\u00e9todo opcional. Retenci\u00f3n menor, posible ISR a cargo en declaraci\u00f3n anual.")
    r = _ins_write_text(ws, r,
        "La columna 'M\u00e9todo Recomendado' indica cu\u00e1l genera mayor PTU neta para el trabajador.")
    r += 1

    r = _ins_write_step(ws, r, 5, "Consultar el resumen (hoja Resumen)")
    r = _ins_write_text(ws, r,
        "Totales de PTU, ISR y netos. Vista r\u00e1pida por empleado con el m\u00e9todo recomendado.")
    r += 1

    r = _ins_write_step(ws, r, 6, "Generar recibos (hoja Recibos)")
    r = _ins_write_text(ws, r,
        "La hoja Recibos contiene una plantilla lista para imprimir. "
        "Si deseas generar PDFs autom\u00e1ticamente, importa el archivo ModuloRecibosPTU.bas "
        "en una copia del archivo guardada como .xlsm (con macros).")
    r += 1

    r = _ins_write_step(ws, r, 7, "Preparar pre-n\u00f3mina para timbrado (hoja Pre_N\u00f3mina)")
    r = _ins_write_text(ws, r,
        "El borrador CFDI muestra el desglose gravado/exento por empleado con claves "
        "de percepci\u00f3n (003-PTU) y deducci\u00f3n (002-ISR). Lleva esta informaci\u00f3n a tu PAC o al portal del SAT.")
    r += 1

    r = _ins_write_separator(ws, r)
    r += 1

    # Comparison table
    r = _ins_write_header(ws, r, "Art. 96 vs Art. 174: Comparaci\u00f3n")
    r = _ins_write_table_row(ws, r, "Aspecto", "Art. 96 LISR (Ley)  |  Art. 174 RLISR (Reglamento)")
    r = _ins_write_table_row(ws, r, "Obligatoriedad", "Obligatorio  |  Opcional")
    r = _ins_write_table_row(ws, r, "Retenci\u00f3n inmediata", "Alta  |  Baja")
    r = _ins_write_table_row(ws, r, "Ajuste anual", "Sin diferencias  |  Posible ISR a cargo")
    r = _ins_write_table_row(ws, r, "Subsidio al empleo", "S\u00ed aplica  |  No aplica")
    r = _ins_write_table_row(ws, r, "Beneficia a", "Patr\u00f3n (sin adeudos)  |  Trabajador (mayor neto)")
    r += 1

    r = _ins_write_separator(ws, r)
    r += 1

    # Legal notes
    r = _ins_write_header(ws, r, "Notas importantes")
    r = _ins_write_note(ws, r,
        "Personas Morales: declaraci\u00f3n anual en marzo, PTU a m\u00e1s tardar el 30 de mayo.")
    r = _ins_write_note(ws, r,
        "Personas F\u00edsicas: declaraci\u00f3n anual en abril, PTU a m\u00e1s tardar el 29 de junio.")
    r = _ins_write_note(ws, r,
        "La PTU no cobrada prescribe en 1 a\u00f1o y se suma al monto del siguiente ejercicio.")
    r = _ins_write_note(ws, r,
        "Trabajadores eventuales requieren m\u00ednimo 60 d\u00edas trabajados para tener derecho.")
    r = _ins_write_note(ws, r,
        "El tope del Art. 127 fr. VIII (reforma 2021) aplica: MAX(3 meses salario, promedio 3 a\u00f1os PTU).")
    r = _ins_write_note(ws, r,
        "Para trabajadores de confianza, el salario tope se calcula autom\u00e1ticamente (120% del m\u00e1s alto de planta).")
    r += 1

    # Footer
    r += 1
    cell = ws.cell(row=r, column=2,
                   value="Columna 13  \u2022  Calculadora de PTU  \u2022  Versi\u00f3n 1.0  \u2022  Marzo 2026")
    cell.font = Font(name=FONT_FAMILY, size=9, color="BDC3C7")

    ws.protection.sheet = True
    ws.protection.enable()


# ==============================================================================
# CONFIG SHEET
# ==============================================================================

def create_config_sheet(wb):
    ws = wb.create_sheet("Config")
    ws.sheet_properties.tabColor = "3498DB"

    set_col_widths(ws, {"A": 28, "B": 20, "C": 50})

    # Title
    ws.cell(row=1, column=1, value="Configuraci\u00f3n").font = FONT_TITLE

    # Parameters
    params = [
        (2,  "Ejercicio",                2025,       "A\u00f1o fiscal para todos los c\u00e1lculos"),
        (3,  "Fecha de Pago",            "30/05/2025", "Fecha l\u00edmite de pago de PTU"),
        (4,  "Tipo de Persona",          "Moral",    "Moral = marzo, F\u00edsica = abril"),
        (5,  "UMA diaria",               None,       "Se calcula autom\u00e1ticamente del ejercicio"),
        (6,  "SMG diario",               None,       "Se calcula autom\u00e1ticamente del ejercicio"),
        (7,  "SMG Frontera",             None,       "Se calcula autom\u00e1ticamente del ejercicio"),
        (8,  "Usar UMA para exenci\u00f3n", True,    "VERDADERO = SAT (UMA), FALSO = PRODECON (SMG)"),
        (9,  "Porcentaje PTU",           0.10,       "10% fijo por Ley (Art. 120 LFT)"),
        (10, "D\u00edas de exenci\u00f3n",  15,      "15 d\u00edas de UMA o SMG (Art. 93 fr. XIV LISR)"),
    ]

    for row_num, label, value, desc in params:
        ws.cell(row=row_num, column=1, value=label).font = FONT_LABEL
        if value is not None:
            cell_b = ws.cell(row=row_num, column=2, value=value)
            cell_b.font = FONT_VALUE
            if row_num == 9:
                cell_b.number_format = '0%'
        ws.cell(row=row_num, column=3, value=desc).font = FONT_SMALL

    # Date format for row 3
    import datetime
    ws.cell(row=3, column=2, value=datetime.date(2025, 5, 30))
    ws["B3"].number_format = "DD/MM/YYYY"

    # Formulas for UMA/SMG (rows 5-7): VLOOKUP into table below
    ws["B5"] = "=IFERROR(VLOOKUP(B2,tbl_UMA_SMG,2,FALSE),\"\")"
    ws["B5"].number_format = '#,##0.00'
    ws["B6"] = "=IFERROR(VLOOKUP(B2,tbl_UMA_SMG,3,FALSE),\"\")"
    ws["B6"].number_format = '#,##0.00'
    ws["B7"] = "=IFERROR(VLOOKUP(B2,tbl_UMA_SMG,4,FALSE),\"\")"
    ws["B7"].number_format = '#,##0.00'

    # Boolean for row 8
    ws["B8"] = True

    # Data validation: Tipo Persona
    dv_tipo = DataValidation(type="list", formula1='"Moral,F\u00edsica"',
                             allow_blank=False, showErrorMessage=True)
    dv_tipo.sqref = "B4"
    ws.add_data_validation(dv_tipo)

    # Data validation: Usar UMA
    dv_uma = DataValidation(type="list", formula1='"VERDADERO,FALSO"',
                            allow_blank=False, showErrorMessage=True)
    dv_uma.sqref = "B8"
    ws.add_data_validation(dv_uma)

    # ── UMA/SMG Table ──────────────────────────────────────────
    tbl_start = 13
    ws.cell(row=12, column=1, value="Tabla UMA / SMG por a\u00f1o").font = FONT_SECTION

    headers = ["A\u00f1o", "UMA Diaria", "SMG Diario", "SMG Frontera"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=tbl_start, column=ci, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER

    for i, (year, uma, smg, smg_f) in enumerate(UMA_SMG_DATA):
        r = tbl_start + 1 + i
        ws.cell(row=r, column=1, value=year).font = FONT_NORMAL
        ws.cell(row=r, column=2, value=uma).number_format = '#,##0.00'
        ws.cell(row=r, column=3, value=smg).number_format = '#,##0.00'
        ws.cell(row=r, column=4, value=smg_f).number_format = '#,##0.00'
        if r % 2 == 0:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = FILL_LIGHT

    last_row = tbl_start + len(UMA_SMG_DATA)

    # Named ranges
    add_named_range(wb, "Ejercicio", attr_text="Config!$B$2")
    add_named_range(wb, "FechaPago", attr_text="Config!$B$3")
    add_named_range(wb, "TipoPersona", attr_text="Config!$B$4")
    add_named_range(wb, "cfg_UMA", attr_text="Config!$B$5")
    add_named_range(wb, "cfg_SMG", attr_text="Config!$B$6")
    add_named_range(wb, "cfg_SMG_Frontera", attr_text="Config!$B$7")
    add_named_range(wb, "UsarUMA", attr_text="Config!$B$8")
    add_named_range(wb, "PctPTU", attr_text="Config!$B$9")
    add_named_range(wb, "DiasExencion", attr_text="Config!$B$10")
    add_named_range(wb, "tbl_UMA_SMG",
                    attr_text=f"Config!$A${tbl_start + 1}:$D${last_row}")

    ws.freeze_panes = "A2"


# ==============================================================================
# TARIFA SHEET
# ==============================================================================

def create_tarifa_sheet(wb):
    ws = wb.create_sheet("Tarifa")
    ws.sheet_properties.tabColor = "E74C3C"

    set_col_widths(ws, {"A": 16, "B": 16, "C": 16, "D": 14})

    # ── Title ──
    ws.cell(row=1, column=1, value="Tarifa Art. 96 LISR \u2014 Mensual").font = FONT_TITLE

    # ── 2024-2025 tariff ──
    ws.cell(row=3, column=1, value="Tarifa 2024-2025").font = FONT_SECTION
    _write_tarifa_block(ws, 4, TARIFA_2024_2025)

    # ── 2026 tariff ──
    ws.cell(row=17, column=1, value="Tarifa 2026").font = FONT_SECTION
    _write_tarifa_block(ws, 18, TARIFA_2026)

    # ── Active tariff (rows 31-42) ── with IF formulas
    ws.cell(row=30, column=1, value="Tarifa Activa (seg\u00fan Ejercicio)").font = FONT_SECTION
    headers = ["L\u00edmite Inferior", "L\u00edmite Superior", "Cuota Fija", "% s/excedente"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=31, column=ci, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER

    for i in range(11):
        r = 32 + i
        r_25 = 5 + i   # 2024-2025 row
        r_26 = 19 + i   # 2026 row
        for c in range(1, 5):
            cl = get_column_letter(c)
            ws[f"{cl}{r}"] = f"=IF(Ejercicio<=2025,{cl}{r_25},{cl}{r_26})"
            if c <= 3:
                ws[f"{cl}{r}"].number_format = '#,##0.00'
            else:
                ws[f"{cl}{r}"].number_format = '0.00%'

    # Named range for active tariff
    add_named_range(wb, "Tarifa96", attr_text="Tarifa!$A$32:$D$42")

    ws.freeze_panes = "A2"


def _write_tarifa_block(ws, header_row, tarifa_data):
    """Write a tariff block with header + 11 data rows."""
    headers = ["L\u00edmite Inferior", "L\u00edmite Superior", "Cuota Fija", "% s/excedente"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER

    for i, (li, ls, cf, pct) in enumerate(tarifa_data):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=li).number_format = '#,##0.00'
        ws.cell(row=r, column=2, value=ls).number_format = '#,##0.00'
        ws.cell(row=r, column=3, value=cf).number_format = '#,##0.00'
        ws.cell(row=r, column=4, value=pct).number_format = '0.00%'
        if r % 2 == 0:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = FILL_LIGHT


# ==============================================================================
# DATOS SHEET
# ==============================================================================

def create_datos_sheet(wb):
    ws = wb.create_sheet("Datos")
    ws.sheet_properties.tabColor = "27AE60"

    # ── Company section (rows 1-10) ──
    ws.cell(row=1, column=1, value="DATOS DE LA EMPRESA Y EMPLEADOS").font = FONT_TITLE

    company_labels = [
        (3,  "Empresa"),
        (4,  "RFC"),
        (5,  "Ejercicio"),
        (6,  "Utilidad Fiscal"),
        (7,  "PTU No Cobrada (ejercicios anteriores)"),
        (8,  "PTU Generada (10%)"),
        (9,  "PTU a Repartir"),
        (10, "50% D\u00edas Trabajados"),
        (11, "50% Percepci\u00f3n Anual"),
    ]
    for row_num, label in company_labels:
        ws.cell(row=row_num, column=1, value=label).font = FONT_LABEL

    # Linked/formula values
    ws["B5"] = "=Ejercicio"
    ws["B6"].number_format = '#,##0.00'
    ws["B7"].number_format = '#,##0.00'

    # Calculated (green bg)
    calc_rows = [8, 9, 10, 11]
    ws["B8"] = "=B6*PctPTU"
    ws["B9"] = "=B8+B7"
    ws["B10"] = "=B9/2"
    ws["B11"] = "=B9/2"

    for row_num in calc_rows:
        ws.cell(row=row_num, column=1).font = FONT_FORMULA_LABEL
        ws.cell(row=row_num, column=2).number_format = '#,##0.00'
        ws.cell(row=row_num, column=2).fill = FILL_SUCCESS_LIGHT

    # ── Column widths ──
    widths = {
        "A": 6,  "B": 28, "C": 15, "D": 20, "E": 14, "F": 14,
        "G": 12, "H": 16, "I": 12, "J": 16, "K": 14, "L": 14,
        "M": 14, "N": 16, "O": 16,
        "P": 12, "Q": 14, "R": 12, "S": 14, "T": 14,
        "U": 14, "V": 14, "W": 14, "X": 14,
    }
    set_col_widths(ws, widths)

    # ── Employee table headers (row 13) ──
    input_headers = [
        ("A", "No."),
        ("B", "Nombre"),
        ("C", "RFC"),
        ("D", "CURP"),
        ("E", "Fecha\nInicio"),
        ("F", "Salario\nDiario"),
        ("G", "D\u00edas\nTrabajados"),
        ("H", "Percepci\u00f3n\nAnual"),
        ("I", "\u00bfConfianza?"),
        ("K", "PTU\nA\u00f1o-3"),
        ("L", "PTU\nA\u00f1o-2"),
        ("M", "PTU\nA\u00f1o-1"),
        ("N", "Ingreso\nMensual"),
        ("O", "ISR Mensual\nOrdinario"),
    ]
    calc_headers = [
        ("J", "Sal. Tope\nConfianza"),
        ("P", "Factor\nD\u00edas"),
        ("Q", "PTU\nD\u00edas"),
        ("R", "Factor\nSalarios"),
        ("S", "PTU\nSalarios"),
        ("T", "PTU\nBruta"),
        ("U", "Tope\n3 Meses"),
        ("V", "Promedio\n3 A\u00f1os"),
        ("W", "Monto\nM\u00e1ximo"),
        ("X", "PTU\nReal"),
    ]

    hr = DAT_HEADER_ROW
    for col_letter, title in input_headers:
        ci = openpyxl.utils.column_index_from_string(col_letter)
        cell = ws.cell(row=hr, column=ci, value=title)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER

    for col_letter, title in calc_headers:
        ci = openpyxl.utils.column_index_from_string(col_letter)
        cell = ws.cell(row=hr, column=ci, value=title)
        cell.font = FONT_CALC_HEADER
        cell.fill = FILL_CALC_HEADER
        cell.alignment = ALIGN_CENTER

    # ── Data validation: Confianza dropdown ──
    dv_conf = DataValidation(type="list", formula1='"S\u00ed,No"',
                             allow_blank=True, showErrorMessage=True,
                             errorTitle="Valor inv\u00e1lido",
                             error="Selecciona S\u00ed o No.")
    dv_conf.sqref = f"I{DAT_DATA_START}:I{DAT_DATA_END}"
    ws.add_data_validation(dv_conf)

    # ── Formulas for each employee row ──
    DS = DAT_DATA_START
    DE = DAT_DATA_END
    for r in range(DS, DE + 1):
        # J: Salario Tope Confianza (auto: MAXIFS * 1.2)
        ws[f"J{r}"] = (
            f'=IF(I{r}="S\u00ed",'
            f'MAXIFS($F${DS}:$F${DE},$I${DS}:$I${DE},"No")*1.2,"")'
        )

        # P: Factor Dias
        ws[f"P{r}"] = (
            f'=IF(G{r}="","",G{r}/SUM($G${DS}:$G${DE}))'
        )

        # Q: PTU Dias
        ws[f"Q{r}"] = (
            f'=IF(P{r}="","",$B$10*P{r})'
        )

        # R: Factor Salarios — adjusts for confianza cap
        ws[f"R{r}"] = (
            f'=IF(H{r}="","",IF(AND(I{r}="S\u00ed",J{r}<>"",F{r}>J{r}),'
            f'MIN(H{r},J{r}*365)/SUM($H${DS}:$H${DE}),'
            f'H{r}/SUM($H${DS}:$H${DE})))'
        )

        # S: PTU Salarios
        ws[f"S{r}"] = (
            f'=IF(R{r}="","",$B$11*R{r})'
        )

        # T: PTU Bruta
        ws[f"T{r}"] = (
            f'=IF(Q{r}="","",Q{r}+S{r})'
        )

        # U: Tope 3 Meses (salario * 91.2)
        ws[f"U{r}"] = (
            f'=IF(F{r}="","",F{r}*91.2)'
        )

        # V: Promedio 3 Años
        ws[f"V{r}"] = (
            f'=IF(F{r}="","",AVERAGE(K{r},L{r},M{r}))'
        )

        # W: Monto Máximo = MAX(tope, promedio)
        ws[f"W{r}"] = (
            f'=IF(U{r}="","",MAX(U{r},V{r}))'
        )

        # X: PTU Real = MIN(bruta, maximo)
        ws[f"X{r}"] = (
            f'=IF(T{r}="","",MIN(T{r},W{r}))'
        )

        # ── Number formats ──
        ws[f"E{r}"].number_format = 'DD/MM/YYYY'
        ws[f"G{r}"].number_format = '0'
        for cl in ["F", "H", "J", "K", "L", "M", "N", "O",
                    "Q", "S", "T", "U", "V", "W", "X"]:
            ws[f"{cl}{r}"].number_format = '#,##0.00'
        for cl in ["P", "R"]:
            ws[f"{cl}{r}"].number_format = '0.000000'

        # ── Alternating row fill ──
        if r % 2 == 0:
            for col in range(1, 25):  # A-X
                ws.cell(row=r, column=col).fill = FILL_LIGHT

    # ── Totals row ──
    tr = DAT_TOTALS_ROW
    ws.cell(row=tr, column=1, value="TOTALES").font = Font(
        name=FONT_FAMILY, bold=True, size=11, color=COLOR_HEADER_BG)
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=5)

    sum_cols = ["G", "H", "Q", "S", "T", "X"]
    for cl in sum_cols:
        ci = openpyxl.utils.column_index_from_string(cl)
        cell = ws.cell(row=tr, column=ci)
        cell.value = f"=SUM({cl}{DS}:{cl}{DE})"
        cell.font = Font(name=FONT_FAMILY, bold=True, size=11)
        cell.number_format = '#,##0.00'
        cell.border = DOUBLE_TOP

    ws.freeze_panes = f"A{DAT_DATA_START}"


# ==============================================================================
# CALCULO_ISR SHEET
# ==============================================================================

def create_calculo_isr_sheet(wb):
    ws = wb.create_sheet("C\u00e1lculo_ISR")
    ws.sheet_properties.tabColor = "E74C3C"

    widths = {
        "A": 6,  "B": 28, "C": 14, "D": 14, "E": 14,
        "F": 14, "G": 14, "H": 14, "I": 14, "J": 14,
        "K": 14, "L": 14, "M": 14, "N": 14, "O": 14,
        "P": 12, "Q": 14, "R": 14,
        "S": 14, "T": 16, "U": 14, "V": 14,
    }
    set_col_widths(ws, widths)

    # ── Headers ──
    hr = ISR_HEADER_ROW

    # Group 1: Base (A-E) navy
    base_headers = [
        ("A", "No."), ("B", "Nombre"), ("C", "PTU\nReal"),
        ("D", "PTU\nExenta"), ("E", "PTU\nGravada"),
    ]
    for cl, title in base_headers:
        ci = openpyxl.utils.column_index_from_string(cl)
        cell = ws.cell(row=hr, column=ci, value=title)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER

    # Group 2: Art. 96 (F-J) navy with accent
    art96_headers = [
        ("F", "Base\nArt.96"), ("G", "ISR Total\nArt.96"),
        ("H", "ISR\nOrdinario"), ("I", "ISR PTU\nArt.96"),
        ("J", "PTU Neta\nArt.96"),
    ]
    fill_96 = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
    for cl, title in art96_headers:
        ci = openpyxl.utils.column_index_from_string(cl)
        cell = ws.cell(row=hr, column=ci, value=title)
        cell.font = FONT_HEADER
        cell.fill = fill_96
        cell.alignment = ALIGN_CENTER

    # Group 3: Art. 174 (K-R) green
    art174_headers = [
        ("K", "PTU Prom\nMensual"), ("L", "Base\nPromediada"),
        ("M", "ISR Base\nProm"), ("N", "ISR Ord\n(sin sub)"),
        ("O", "Diferencia\nISR"), ("P", "Tasa\nEfectiva"),
        ("Q", "ISR PTU\nArt.174"), ("R", "PTU Neta\nArt.174"),
    ]
    for cl, title in art174_headers:
        ci = openpyxl.utils.column_index_from_string(cl)
        cell = ws.cell(row=hr, column=ci, value=title)
        cell.font = FONT_CALC_HEADER
        cell.fill = FILL_CALC_HEADER
        cell.alignment = ALIGN_CENTER

    # Group 4: Comparison (S-V) navy
    comp_headers = [
        ("S", "Diferencia\nISR"), ("T", "M\u00e9todo\nRecomendado"),
        ("U", "ISR\nRecomendado"), ("V", "PTU Neta\nFinal"),
    ]
    for cl, title in comp_headers:
        ci = openpyxl.utils.column_index_from_string(cl)
        cell = ws.cell(row=hr, column=ci, value=title)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER

    # ── ISR VLOOKUP formula helper ──
    # ISR = (base - VLOOKUP(base, Tarifa96, 1, TRUE)) * VLOOKUP(base, Tarifa96, 4, TRUE)
    #       + VLOOKUP(base, Tarifa96, 3, TRUE)
    def isr_formula(base_cell):
        return (
            f"({base_cell}-VLOOKUP({base_cell},Tarifa96,1,TRUE))"
            f"*VLOOKUP({base_cell},Tarifa96,4,TRUE)"
            f"+VLOOKUP({base_cell},Tarifa96,3,TRUE)"
        )

    # ── Formulas for each row ──
    for i in range(NUM_EMPLOYEES):
        r = ISR_DATA_START + i
        dr = DAT_DATA_START + i  # corresponding Datos row

        # A: No
        ws[f"A{r}"] = f"=Datos!A{dr}"
        # B: Nombre
        ws[f"B{r}"] = f"=Datos!B{dr}"
        # C: PTU Real
        ws[f"C{r}"] = f"=Datos!X{dr}"

        # D: PTU Exenta = IF(UsarUMA, cfg_UMA, cfg_SMG) * DiasExencion
        ws[f"D{r}"] = (
            f'=IF(C{r}="","",IF(UsarUMA,cfg_UMA,cfg_SMG)*DiasExencion)'
        )

        # E: PTU Gravada = MAX(0, Real - Exenta)
        ws[f"E{r}"] = f'=IF(C{r}="","",MAX(0,C{r}-D{r}))'

        # ── Art. 96 ──
        # F: Base Art.96 = ingreso mensual + PTU gravada
        ws[f"F{r}"] = f'=IF(E{r}="","",Datos!N{dr}+E{r})'

        # G: ISR Total Art.96
        ws[f"G{r}"] = f'=IF(F{r}="","",IFERROR({isr_formula(f"F{r}")},0))'

        # H: ISR Ordinario (user-entered from Datos)
        ws[f"H{r}"] = f'=IF(C{r}="","",Datos!O{dr})'

        # I: ISR PTU Art.96 = ISR total - ISR ordinario
        ws[f"I{r}"] = f'=IF(G{r}="","",MAX(0,G{r}-H{r}))'

        # J: PTU Neta Art.96
        ws[f"J{r}"] = f'=IF(C{r}="","",C{r}-I{r})'

        # ── Art. 174 ──
        # K: PTU Promedio Mensual = (gravada/365)*30.4
        ws[f"K{r}"] = f'=IF(E{r}="","",(E{r}/365)*30.4)'

        # L: Base Promediada = ingreso mensual + PTU prom
        ws[f"L{r}"] = f'=IF(K{r}="","",Datos!N{dr}+K{r})'

        # M: ISR Base Promediada
        ws[f"M{r}"] = f'=IF(L{r}="","",IFERROR({isr_formula(f"L{r}")},0))'

        # N: ISR Ordinario (sin subsidio) — recalculated via tarifa
        ws[f"N{r}"] = (
            f'=IF(C{r}="","",IFERROR({isr_formula(f"Datos!N{dr}")},0))'
        )

        # O: Diferencia ISR
        ws[f"O{r}"] = f'=IF(M{r}="","",M{r}-N{r})'

        # P: Tasa Efectiva
        ws[f"P{r}"] = f'=IF(K{r}="","",IF(K{r}>0,O{r}/K{r},0))'

        # Q: ISR PTU Art.174
        ws[f"Q{r}"] = f'=IF(E{r}="","",E{r}*P{r})'

        # R: PTU Neta Art.174
        ws[f"R{r}"] = f'=IF(C{r}="","",C{r}-Q{r})'

        # ── Comparison ──
        # S: Diferencia
        ws[f"S{r}"] = f'=IF(I{r}="","",ABS(I{r}-Q{r}))'

        # T: Método Recomendado
        ws[f"T{r}"] = f'=IF(C{r}="","",IF(Q{r}<I{r},"Art. 174","Art. 96"))'

        # U: ISR Recomendado
        ws[f"U{r}"] = f'=IF(C{r}="","",MIN(I{r},Q{r}))'

        # V: PTU Neta Final
        ws[f"V{r}"] = f'=IF(C{r}="","",C{r}-U{r})'

        # ── Number formats ──
        for cl in ["C", "D", "E", "F", "G", "H", "I", "J",
                    "K", "L", "M", "N", "O", "Q", "R",
                    "S", "U", "V"]:
            ws[f"{cl}{r}"].number_format = '#,##0.00'
        ws[f"P{r}"].number_format = '0.0000%'

        # Alternating rows
        if r % 2 == 0:
            for col in range(1, 23):
                ws.cell(row=r, column=col).fill = FILL_LIGHT

    # ── Totals row ──
    tr = ISR_DATA_END + 1
    ws.cell(row=tr, column=1, value="TOTALES").font = Font(
        name=FONT_FAMILY, bold=True, size=11, color=COLOR_HEADER_BG)
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=2)

    for cl in ["C", "D", "E", "I", "J", "Q", "R", "S", "U", "V"]:
        ci = openpyxl.utils.column_index_from_string(cl)
        cell = ws.cell(row=tr, column=ci)
        cell.value = f"=SUM({cl}{ISR_DATA_START}:{cl}{ISR_DATA_END})"
        cell.font = Font(name=FONT_FAMILY, bold=True, size=11)
        cell.number_format = '#,##0.00'
        cell.border = DOUBLE_TOP

    ws.freeze_panes = f"C{ISR_DATA_START}"


# ==============================================================================
# RESUMEN SHEET
# ==============================================================================

def create_resumen_sheet(wb):
    ws = wb.create_sheet("Resumen")
    ws.sheet_properties.tabColor = "8E44AD"

    set_col_widths(ws, {
        "A": 6, "B": 28, "C": 16, "D": 16, "E": 16,
        "F": 18, "G": 16,
    })

    ws.cell(row=1, column=1, value="Resumen de PTU").font = FONT_TITLE
    ws.cell(row=2, column=1, value="Ejercicio:").font = FONT_LABEL
    ws["B2"] = "=Ejercicio"

    # ── Totals section ──
    ws.cell(row=4, column=1, value="Totales Generales").font = FONT_SECTION

    summary_labels = [
        (5, "PTU Generada (10%)",    "=Datos!B8"),
        (6, "PTU No Cobrada",         "=Datos!B7"),
        (7, "PTU a Repartir",         "=Datos!B9"),
        (8, "Total PTU Real (despu\u00e9s de topes)",
             f"=SUM(Datos!X{DAT_DATA_START}:X{DAT_DATA_END})"),
        (9, "Total ISR Art. 96",
             f"=SUM(C\u00e1lculo_ISR!I{ISR_DATA_START}:I{ISR_DATA_END})"),
        (10, "Total ISR Art. 174",
             f"=SUM(C\u00e1lculo_ISR!Q{ISR_DATA_START}:Q{ISR_DATA_END})"),
        (11, "Total ISR Recomendado",
             f"=SUM(C\u00e1lculo_ISR!U{ISR_DATA_START}:U{ISR_DATA_END})"),
        (12, "Total PTU Neta a Pagar",
             f"=SUM(C\u00e1lculo_ISR!V{ISR_DATA_START}:V{ISR_DATA_END})"),
    ]
    for row_num, label, formula in summary_labels:
        ws.cell(row=row_num, column=1, value=label).font = FONT_LABEL
        ws.cell(row=row_num, column=2, value=formula).number_format = '#,##0.00'
        if row_num >= 8:
            ws.cell(row=row_num, column=2).fill = FILL_SUCCESS_LIGHT

    # ── Per-employee table ──
    ws.cell(row=14, column=1, value="Detalle por Empleado").font = FONT_SECTION

    emp_headers = [
        ("A", "No."), ("B", "Nombre"), ("C", "PTU Real"),
        ("D", "ISR Art.96"), ("E", "ISR Art.174"),
        ("F", "M\u00e9todo"), ("G", "PTU Neta"),
    ]
    for cl, title in emp_headers:
        ci = openpyxl.utils.column_index_from_string(cl)
        cell = ws.cell(row=15, column=ci, value=title)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER

    for i in range(NUM_EMPLOYEES):
        r = 16 + i
        ir = ISR_DATA_START + i  # Calculo_ISR row

        ws[f"A{r}"] = f"=C\u00e1lculo_ISR!A{ir}"
        ws[f"B{r}"] = f"=C\u00e1lculo_ISR!B{ir}"
        ws[f"C{r}"] = f"=C\u00e1lculo_ISR!C{ir}"
        ws[f"D{r}"] = f"=C\u00e1lculo_ISR!I{ir}"
        ws[f"E{r}"] = f"=C\u00e1lculo_ISR!Q{ir}"
        ws[f"F{r}"] = f"=C\u00e1lculo_ISR!T{ir}"
        ws[f"G{r}"] = f"=C\u00e1lculo_ISR!V{ir}"

        for cl in ["C", "D", "E", "G"]:
            ws[f"{cl}{r}"].number_format = '#,##0.00'

        if r % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=r, column=col).fill = FILL_LIGHT

    # Totals
    tr = 16 + NUM_EMPLOYEES
    ws.cell(row=tr, column=1, value="TOTALES").font = Font(
        name=FONT_FAMILY, bold=True, size=11, color=COLOR_HEADER_BG)
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=2)
    for cl in ["C", "D", "E", "G"]:
        ci = openpyxl.utils.column_index_from_string(cl)
        cell = ws.cell(row=tr, column=ci)
        cell.value = f"=SUM({cl}16:{cl}{tr - 1})"
        cell.font = Font(name=FONT_FAMILY, bold=True, size=11)
        cell.number_format = '#,##0.00'
        cell.border = DOUBLE_TOP

    ws.freeze_panes = "A16"


# ==============================================================================
# RECIBOS SHEET
# ==============================================================================

def create_recibos_sheet(wb):
    ws = wb.create_sheet("Recibos")
    ws.sheet_properties.tabColor = "F39C12"

    set_col_widths(ws, {
        "A": 4, "B": 25, "C": 20, "D": 20, "E": 20, "F": 4,
    })

    # ── Receipt template (references employee #1 as preview) ──
    dr = DAT_DATA_START  # first employee in Datos
    ir = ISR_DATA_START  # first employee in Calculo_ISR

    # Company header
    ws.merge_cells("B2:E2")
    ws["B2"] = "=Datos!B3"  # Empresa
    ws["B2"].font = Font(name=FONT_FAMILY, bold=True, size=16, color=COLOR_HEADER_BG)
    ws["B2"].alignment = ALIGN_CENTER

    ws.merge_cells("B3:E3")
    ws["B3"] = "=Datos!B4"  # RFC
    ws["B3"].font = Font(name=FONT_FAMILY, size=11, color=COLOR_MUTED)
    ws["B3"].alignment = ALIGN_CENTER

    ws.merge_cells("B5:E5")
    ws["B5"] = "=CONCATENATE(\"RECIBO DE PTU \u2014 Ejercicio \",Ejercicio)"
    ws["B5"].font = Font(name=FONT_FAMILY, bold=True, size=14, color=COLOR_HEADER_BG)
    ws["B5"].alignment = ALIGN_CENTER

    # Employee data
    emp_fields = [
        (7,  "Trabajador:", f"=Datos!B{dr}"),
        (8,  "RFC:",        f"=Datos!C{dr}"),
        (9,  "CURP:",       f"=Datos!D{dr}"),
        (10, "Fecha de pago:", "=FechaPago"),
    ]
    for row_num, label, formula in emp_fields:
        ws.cell(row=row_num, column=2, value=label).font = FONT_LABEL
        ws.cell(row=row_num, column=3, value=formula).font = FONT_NORMAL
    ws[f"C10"].number_format = "DD/MM/YYYY"

    # Breakdown table
    ws.cell(row=12, column=2, value="Concepto").font = FONT_HEADER
    ws.cell(row=12, column=2).fill = FILL_HEADER
    ws.cell(row=12, column=3, value="Importe").font = FONT_HEADER
    ws.cell(row=12, column=3).fill = FILL_HEADER

    breakdown = [
        (13, "PTU Bruta",         f"=Datos!T{dr}"),
        (14, "Tope aplicado",     f"=Datos!W{dr}"),
        (15, "PTU Real",          f"=Datos!X{dr}"),
        (16, "PTU Exenta",        f"=C\u00e1lculo_ISR!D{ir}"),
        (17, "PTU Gravada",       f"=C\u00e1lculo_ISR!E{ir}"),
        (18, "ISR Retenido",      f"=C\u00e1lculo_ISR!U{ir}"),
        (19, "M\u00e9todo ISR",   f"=C\u00e1lculo_ISR!T{ir}"),
    ]
    for row_num, label, formula in breakdown:
        ws.cell(row=row_num, column=2, value=label).font = FONT_NORMAL
        ws.cell(row=row_num, column=3, value=formula)
        if row_num != 19:
            ws[f"C{row_num}"].number_format = '#,##0.00'
        if row_num % 2 == 0:
            ws.cell(row=row_num, column=2).fill = FILL_LIGHT
            ws.cell(row=row_num, column=3).fill = FILL_LIGHT

    # Net pay highlight
    ws.cell(row=21, column=2, value="PTU NETA A RECIBIR").font = Font(
        name=FONT_FAMILY, bold=True, size=14, color=COLOR_SUCCESS)
    ws.cell(row=21, column=3, value=f"=C\u00e1lculo_ISR!V{ir}")
    ws[f"C21"].number_format = '#,##0.00'
    ws[f"C21"].font = Font(name=FONT_FAMILY, bold=True, size=14, color=COLOR_SUCCESS)
    ws.cell(row=21, column=2).fill = FILL_SUCCESS_LIGHT
    ws.cell(row=21, column=3).fill = FILL_SUCCESS_LIGHT

    # Signature block
    ws.merge_cells("B24:E24")
    ws["B24"] = "Recib\u00ed de conformidad la cantidad arriba se\u00f1alada."
    ws["B24"].font = Font(name=FONT_FAMILY, size=11, color=COLOR_MUTED)
    ws["B24"].alignment = ALIGN_CENTER

    ws.merge_cells("B27:D27")
    ws["B27"] = "_" * 50
    ws["B27"].alignment = ALIGN_CENTER

    ws.merge_cells("B28:D28")
    ws["B28"] = "Nombre y firma del trabajador"
    ws["B28"].font = Font(name=FONT_FAMILY, size=10, color=COLOR_MUTED)
    ws["B28"].alignment = ALIGN_CENTER

    ws.merge_cells("B30:E30")
    ws["B30"] = ("De conformidad con los art\u00edculos 117 al 131 "
                 "de la Ley Federal del Trabajo.")
    ws["B30"].font = Font(name=FONT_FAMILY, size=8, italic=True, color=COLOR_MUTED)
    ws["B30"].alignment = ALIGN_CENTER

    # Print setup
    ws.print_area = "A1:F31"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = "portrait"


# ==============================================================================
# PRE_NOMINA SHEET
# ==============================================================================

def create_prenomina_sheet(wb):
    ws = wb.create_sheet("Pre_N\u00f3mina")
    ws.sheet_properties.tabColor = "1ABC9C"

    widths = {
        "A": 6, "B": 28, "C": 15, "D": 20, "E": 16,
        "F": 18, "G": 16, "H": 8, "I": 8,
        "J": 14, "K": 14, "L": 14, "M": 14,
        "N": 8, "O": 14, "P": 14, "Q": 14, "R": 16,
    }
    set_col_widths(ws, widths)

    # Title
    ws.cell(row=1, column=1,
            value="Pre-N\u00f3mina PTU \u2014 Borrador para Timbrado CFDI").font = FONT_TITLE
    ws.merge_cells("A1:R1")

    # Headers (row 3)
    hr = 3
    headers = [
        ("A", "No."),
        ("B", "Nombre"),
        ("C", "RFC"),
        ("D", "CURP"),
        ("E", "NSS"),
        ("F", "R\u00e9gimen"),
        ("G", "Tipo N\u00f3mina"),
        ("H", "Clave\nPerc."),
        ("I", "Concepto"),
        ("J", "PTU\nGravada"),
        ("K", "PTU\nExenta"),
        ("L", "Total Perc.\nGravadas"),
        ("M", "Total Perc.\nExentas"),
        ("N", "Clave\nDed."),
        ("O", "ISR\nRetenido"),
        ("P", "Total\nDeducciones"),
        ("Q", "Neto\na Pagar"),
        ("R", "M\u00e9todo\nISR"),
    ]
    for cl, title in headers:
        ci = openpyxl.utils.column_index_from_string(cl)
        cell = ws.cell(row=hr, column=ci, value=title)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER

    # Data rows
    data_start = 4
    for i in range(NUM_EMPLOYEES):
        r = data_start + i
        dr = DAT_DATA_START + i
        ir = ISR_DATA_START + i

        ws[f"A{r}"] = f"=Datos!A{dr}"
        ws[f"B{r}"] = f"=Datos!B{dr}"
        ws[f"C{r}"] = f"=Datos!C{dr}"
        ws[f"D{r}"] = f"=Datos!D{dr}"
        # E: NSS (user input)
        ws[f"F{r}"] = "Sueldos y Salarios"
        ws[f"F{r}"].font = FONT_SMALL
        ws[f"G{r}"] = "Extraordinaria"
        ws[f"G{r}"].font = FONT_SMALL
        ws[f"H{r}"] = "003"
        ws[f"H{r}"].alignment = ALIGN_CENTER
        ws[f"I{r}"] = "PTU"
        ws[f"I{r}"].alignment = ALIGN_CENTER

        # Amounts from Calculo_ISR
        ws[f"J{r}"] = f"=C\u00e1lculo_ISR!E{ir}"  # PTU Gravada
        ws[f"K{r}"] = f"=C\u00e1lculo_ISR!D{ir}"  # PTU Exenta
        ws[f"L{r}"] = f"=J{r}"                      # Total gravadas
        ws[f"M{r}"] = f"=K{r}"                      # Total exentas
        ws[f"N{r}"] = "002"
        ws[f"N{r}"].alignment = ALIGN_CENTER
        ws[f"O{r}"] = f"=C\u00e1lculo_ISR!U{ir}"  # ISR recomendado
        ws[f"P{r}"] = f"=O{r}"                      # Total deducciones
        ws[f"Q{r}"] = f"=C\u00e1lculo_ISR!V{ir}"  # Neto
        ws[f"R{r}"] = f"=C\u00e1lculo_ISR!T{ir}"  # Método

        for cl in ["J", "K", "L", "M", "O", "P", "Q"]:
            ws[f"{cl}{r}"].number_format = '#,##0.00'

        if r % 2 == 0:
            for col in range(1, 19):
                ws.cell(row=r, column=col).fill = FILL_LIGHT

    # Totals
    tr = data_start + NUM_EMPLOYEES
    ws.cell(row=tr, column=1, value="TOTALES").font = Font(
        name=FONT_FAMILY, bold=True, size=11, color=COLOR_HEADER_BG)
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=4)

    for cl in ["J", "K", "L", "M", "O", "P", "Q"]:
        ci = openpyxl.utils.column_index_from_string(cl)
        cell = ws.cell(row=tr, column=ci)
        cell.value = f"=SUM({cl}{data_start}:{cl}{tr - 1})"
        cell.font = Font(name=FONT_FAMILY, bold=True, size=11)
        cell.number_format = '#,##0.00'
        cell.border = DOUBLE_TOP

    ws.freeze_panes = f"E{data_start}"
    ws.print_area = f"A1:R{tr}"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = "landscape"


# ==============================================================================
# MAIN
# ==============================================================================

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    create_instrucciones_sheet(wb)
    create_config_sheet(wb)
    create_tarifa_sheet(wb)
    create_datos_sheet(wb)
    create_calculo_isr_sheet(wb)
    create_resumen_sheet(wb)
    create_recibos_sheet(wb)
    create_prenomina_sheet(wb)

    # Set Instrucciones as active sheet
    wb.active = wb.sheetnames.index("Instrucciones")

    wb.save(str(OUTPUT_FILE))
    print(f"Plantilla generada: {OUTPUT_FILE}")
    print(f"  Hojas: {', '.join(wb.sheetnames)}")
    print(f"  Empleados soportados: {NUM_EMPLOYEES}")


if __name__ == "__main__":
    main()
