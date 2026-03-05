"""
Extractor de Datos de Reportes de Gasolineras
=============================================
Extrae información de archivos Excel con reportes diarios y genera
archivos consolidados con resumen y detalles.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import pandas as pd
from pathlib import Path
from datetime import datetime
from calendar import monthrange
import shutil
import tempfile

# Configuración
INPUT_DIR = Path(__file__).parent / "input"
OUTPUT_DIR = Path(__file__).parent / "output"

MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}


def parsear_nombre_archivo(nombre: str) -> tuple:
    """
    Extrae mes, sucursal y año del nombre de archivo.
    Formato esperado: REPORTE-MES-SUCURSAL-AÑO.xlsx (separadores: '-' o espacio)

    Returns: (mes_nombre, mes_num, sucursal, anio) o None si no se puede parsear.
    """
    base = Path(nombre).stem.upper().replace(" ", "-")
    partes = [p for p in base.split("-") if p and p != "REPORTE"]

    mes_nombre = None
    mes_num = None
    anio = None
    resto = []

    for parte in partes:
        if parte in MESES and mes_nombre is None:
            mes_nombre = parte
            mes_num = MESES[parte]
        elif parte.isdigit() and len(parte) == 4 and anio is None:
            anio = int(parte)
        else:
            resto.append(parte)

    if not mes_nombre or not anio or not resto:
        return None

    sucursal = "_".join(resto)
    return (mes_nombre, mes_num, sucursal, anio)


def copiar_archivo_temp(archivo_original: Path) -> Path:
    """Copia el archivo a temp para evitar conflictos con OneDrive/Excel abierto"""
    temp_dir = Path(tempfile.gettempdir())
    temp_file = temp_dir / f"temp_{archivo_original.name}"
    shutil.copy2(archivo_original, temp_file)
    return temp_file


def extraer_tabla_vertical(sheet, fila_inicio: int, col_inicio: int, cols: list) -> list:
    """
    Extrae una tabla vertical que termina cuando encuentra 'TOTAL' en columna J.
    
    Returns: Lista de diccionarios con los datos de cada fila.
    """
    datos = []
    fila = fila_inicio
    
    while fila <= sheet.max_row:
        # Verificar si llegamos a TOTAL
        valor_col_j = sheet.cell(row=fila, column=10).value  # J=10
        if valor_col_j and "TOTAL" in str(valor_col_j).upper():
            break
        
        # Leer valores de las columnas especificadas
        fila_datos = {}
        tiene_datos = False
        
        for i, col_name in enumerate(cols):
            valor = sheet.cell(row=fila, column=col_inicio + i).value
            fila_datos[col_name] = valor
            if valor is not None:
                tiene_datos = True
        
        # Verificar que FOLIO existe y que IMPORTE es numérico (no es encabezado)
        folio = fila_datos.get('FOLIO')
        importe = fila_datos.get('IMPORTE')
        
        # Saltar filas de encabezado (donde FOLIO='FOLIO' o IMPORTE='IMPORTE')
        es_encabezado = (str(folio).upper() == 'FOLIO' if folio else False) or \
                        (str(importe).upper() == 'IMPORTE' if importe else False)
        
        if tiene_datos and folio and not es_encabezado:
            # Convertir IMPORTE a número si es posible
            try:
                if importe is not None:
                    fila_datos['IMPORTE'] = float(importe)
            except (ValueError, TypeError):
                fila_datos['IMPORTE'] = 0
            datos.append(fila_datos)
        
        fila += 1
    
    return datos


def extraer_gastos(sheet, fila_inicio: int) -> list:
    """Extrae la tabla de gastos desde columna M y N"""
    datos = []
    fila = fila_inicio
    
    while fila <= sheet.max_row:
        concepto = sheet.cell(row=fila, column=13).value  # M=13
        importe = sheet.cell(row=fila, column=14).value   # N=14
        
        if concepto and "TOTAL" in str(concepto).upper():
            break
        
        # Saltar encabezados
        if concepto and str(concepto).upper() in ['CONCEPTO', 'GASTOS']:
            fila += 1
            continue
        
        if concepto and importe:
            try:
                importe_num = float(importe)
            except (ValueError, TypeError):
                importe_num = 0
            
            datos.append({
                'CONCEPTO': concepto,
                'IMPORTE': importe_num
            })
        
        fila += 1
    
    return datos


def extraer_depositos(sheet) -> float:
    """Extrae la suma de DEP 1 y DEP 2 desde O10:P11"""
    dep1 = sheet.cell(row=10, column=16).value or 0  # P10
    dep2 = sheet.cell(row=11, column=16).value or 0  # P11
    
    # Convertir a número
    try:
        dep1 = float(dep1) if dep1 else 0
    except (ValueError, TypeError):
        dep1 = 0
    
    try:
        dep2 = float(dep2) if dep2 else 0
    except (ValueError, TypeError):
        dep2 = 0
    
    return dep1 + dep2


def buscar_seccion(sheet, nombre_seccion: str, col_busqueda: int = 9) -> int:
    """
    Busca una sección por nombre en la columna especificada.
    
    Args:
        sheet: Hoja de Excel
        nombre_seccion: Nombre de la sección a buscar (ej: "VALES", "TARJETA DE CREDITO")
        col_busqueda: Columna donde buscar (default I=9)
    
    Returns:
        Número de fila donde inician los DATOS (después del encabezado), o 0 si no se encuentra
    """
    for fila in range(1, min(sheet.max_row + 1, 60)):  # Buscar en primeras 60 filas
        valor = sheet.cell(row=fila, column=col_busqueda).value
        if valor and nombre_seccion.upper() in str(valor).upper():
            # Encontramos el título, los datos empiezan 2 filas después (después del encabezado FOLIO/CLIENTE/IMPORTE)
            return fila + 2
    
    return 0  # No encontrada


def procesar_hoja(sheet, dia: int, mes: int, anio: int) -> dict:
    """
    Procesa una hoja y extrae toda la información.
    Detecta dinámicamente la posición de VALES y TARJETA DE CREDITO.
    
    Returns: Diccionario con los datos extraídos.
    """
    fecha = datetime(anio, mes, dia)
    
    # Extraer NOTAS DE CREDITO (columna I=9, siempre empieza fila 3)
    notas_credito = extraer_tabla_vertical(sheet, 3, 9, ['FOLIO', 'CLIENTE', 'IMPORTE'])
    total_clientes = sum(n.get('IMPORTE', 0) or 0 for n in notas_credito)
    
    # Buscar dinámicamente VALES (puede cambiar de posición)
    fila_vales = buscar_seccion(sheet, "VALES", 9)
    if fila_vales > 0:
        vales = extraer_tabla_vertical(sheet, fila_vales, 9, ['FOLIO', 'CLIENTE', 'IMPORTE'])
    else:
        vales = []
    total_vales = sum(v.get('IMPORTE', 0) or 0 for v in vales)
    
    # Buscar dinámicamente TARJETA DE CREDITO (puede cambiar de posición)
    fila_tarjetas = buscar_seccion(sheet, "TARJETA DE CREDITO", 9)
    if fila_tarjetas > 0:
        tarjetas = extraer_tabla_vertical(sheet, fila_tarjetas, 9, ['FOLIO', 'CLIENTE', 'IMPORTE'])
    else:
        tarjetas = []
    total_tarjetas = sum(t.get('IMPORTE', 0) or 0 for t in tarjetas)
    
    # Extraer GASTOS (columna M=13, buscar dinámicamente también)
    fila_gastos = buscar_seccion(sheet, "GASTOS", 13)
    if fila_gastos > 0:
        gastos = extraer_gastos(sheet, fila_gastos)
    else:
        gastos = []
    total_gastos = sum(g.get('IMPORTE', 0) or 0 for g in gastos)
    
    # Extraer DEPOSITOS
    total_bancos = extraer_depositos(sheet)
    
    return {
        'fecha': fecha,
        'notas_credito': notas_credito,
        'vales': vales,
        'tarjetas': tarjetas,
        'gastos': gastos,
        'totales': {
            'CLIENTES': total_clientes,
            'VALES': total_vales,
            'TARJETAS': total_tarjetas,
            'GASTOS': total_gastos,
            'BANCOS': total_bancos
        }
    }


def aplicar_estilos(ws, num_filas: int, num_cols: int):
    """Aplica estilos a una hoja de Excel"""
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Aplicar a encabezados (fila 1)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Aplicar bordes a todas las celdas
    for row in range(2, num_filas + 1):
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).border = thin_border
    
    # Ajustar ancho de columnas
    for col in range(1, num_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15


def generar_archivo_consolidado(datos_dias: list, archivo_salida: Path):
    """
    Genera el archivo Excel consolidado con todas las hojas.
    """
    wb = openpyxl.Workbook()
    
    # 1. HOJA RESUMEN
    ws_resumen = wb.active
    ws_resumen.title = "RESUMEN"
    
    # Encabezados
    encabezados_resumen = ['DIA', 'CLIENTES', 'VALES', 'TARJETAS', 'GASTOS', 'BANCOS']
    for col, header in enumerate(encabezados_resumen, 1):
        ws_resumen.cell(row=1, column=col, value=header)
    
    # Datos
    for i, dia_data in enumerate(datos_dias, 2):
        ws_resumen.cell(row=i, column=1, value=dia_data['fecha'].strftime('%d/%m/%Y'))
        ws_resumen.cell(row=i, column=2, value=dia_data['totales']['CLIENTES'])
        ws_resumen.cell(row=i, column=3, value=dia_data['totales']['VALES'])
        ws_resumen.cell(row=i, column=4, value=dia_data['totales']['TARJETAS'])
        ws_resumen.cell(row=i, column=5, value=dia_data['totales']['GASTOS'])
        ws_resumen.cell(row=i, column=6, value=dia_data['totales']['BANCOS'])
    
    aplicar_estilos(ws_resumen, len(datos_dias) + 1, 6)
    
    # Formato de moneda para columnas numéricas
    for row in range(2, len(datos_dias) + 2):
        for col in range(2, 7):
            ws_resumen.cell(row=row, column=col).number_format = '$#,##0.00'
    
    # 2. HOJA CLIENTES (detalle notas de crédito)
    ws_clientes = wb.create_sheet("CLIENTES")
    encabezados_clientes = ['DIA', 'FOLIO', 'CLIENTE', 'IMPORTE']
    for col, header in enumerate(encabezados_clientes, 1):
        ws_clientes.cell(row=1, column=col, value=header)
    
    fila = 2
    for dia_data in datos_dias:
        for nota in dia_data['notas_credito']:
            ws_clientes.cell(row=fila, column=1, value=dia_data['fecha'].strftime('%d/%m/%Y'))
            ws_clientes.cell(row=fila, column=2, value=nota.get('FOLIO'))
            ws_clientes.cell(row=fila, column=3, value=nota.get('CLIENTE'))
            ws_clientes.cell(row=fila, column=4, value=nota.get('IMPORTE'))
            fila += 1
    
    if fila > 2:
        aplicar_estilos(ws_clientes, fila - 1, 4)
        for row in range(2, fila):
            ws_clientes.cell(row=row, column=4).number_format = '$#,##0.00'
    
    # 3. HOJA VALES (detalle)
    ws_vales = wb.create_sheet("VALES")
    encabezados_vales = ['DIA', 'FOLIO', 'CLIENTE', 'IMPORTE']
    for col, header in enumerate(encabezados_vales, 1):
        ws_vales.cell(row=1, column=col, value=header)
    
    fila = 2
    for dia_data in datos_dias:
        for vale in dia_data['vales']:
            ws_vales.cell(row=fila, column=1, value=dia_data['fecha'].strftime('%d/%m/%Y'))
            ws_vales.cell(row=fila, column=2, value=vale.get('FOLIO'))
            ws_vales.cell(row=fila, column=3, value=vale.get('CLIENTE'))
            ws_vales.cell(row=fila, column=4, value=vale.get('IMPORTE'))
            fila += 1
    
    if fila > 2:
        aplicar_estilos(ws_vales, fila - 1, 4)
        for row in range(2, fila):
            ws_vales.cell(row=row, column=4).number_format = '$#,##0.00'
    
    # 4. HOJA GASTOS (detalle)
    ws_gastos = wb.create_sheet("GASTOS")
    encabezados_gastos = ['DIA', 'CONCEPTO', 'IMPORTE']
    for col, header in enumerate(encabezados_gastos, 1):
        ws_gastos.cell(row=1, column=col, value=header)
    
    fila = 2
    for dia_data in datos_dias:
        for gasto in dia_data['gastos']:
            ws_gastos.cell(row=fila, column=1, value=dia_data['fecha'].strftime('%d/%m/%Y'))
            ws_gastos.cell(row=fila, column=2, value=gasto.get('CONCEPTO'))
            ws_gastos.cell(row=fila, column=3, value=gasto.get('IMPORTE'))
            fila += 1
    
    if fila > 2:
        aplicar_estilos(ws_gastos, fila - 1, 3)
        for row in range(2, fila):
            ws_gastos.cell(row=row, column=3).number_format = '$#,##0.00'
    
    # Guardar archivo
    wb.save(archivo_salida)
    print(f"  -> Archivo generado: {archivo_salida}")


def procesar_archivo(nombre_archivo: str, sucursal: str, mes_nombre: str, mes_num: int, anio: int):
    """Procesa un archivo completo y genera el consolidado."""
    archivo_path = INPUT_DIR / nombre_archivo

    if not archivo_path.exists():
        print(f"ERROR: No se encontró el archivo {archivo_path}")
        return

    print(f"\nProcesando: {nombre_archivo}")

    # Copiar a temp para evitar conflictos
    temp_file = copiar_archivo_temp(archivo_path)
    print(f"  Archivo copiado a: {temp_file}")

    try:
        wb = openpyxl.load_workbook(temp_file, data_only=True)

        datos_dias = []

        # Procesar cada hoja según los días válidos del mes
        _, dias_en_mes = monthrange(anio, mes_num)
        for dia in range(1, dias_en_mes + 1):
            # Buscar la hoja por nombre (puede ser "01", "1", etc.)
            nombre_hoja = None
            for sheet_name in wb.sheetnames:
                if sheet_name.strip() in [str(dia), str(dia).zfill(2)]:
                    nombre_hoja = sheet_name
                    break

            if nombre_hoja:
                print(f"  Procesando día {dia}...")
                sheet = wb[nombre_hoja]
                datos = procesar_hoja(sheet, dia, mes_num, anio)
                datos_dias.append(datos)
            else:
                print(f"  Alerta: No se encontró hoja para día {dia}")

        wb.close()

        # Generar archivo consolidado
        OUTPUT_DIR.mkdir(exist_ok=True)
        archivo_salida = OUTPUT_DIR / f"CONSOLIDADO_{sucursal}_{mes_nombre}_{anio}.xlsx"
        generar_archivo_consolidado(datos_dias, archivo_salida)
        
        print(f"\n  Resumen {sucursal}:")
        print(f"    Días procesados: {len(datos_dias)}")
        
    finally:
        # Limpiar archivo temporal
        if temp_file.exists():
            temp_file.unlink()


def main():
    print("=" * 60)
    print("EXTRACTOR DE REPORTES DE GASOLINERAS")
    print("=" * 60)

    archivos = sorted(INPUT_DIR.glob("*.xlsx"))
    if not archivos:
        print(f"\nNo se encontraron archivos .xlsx en {INPUT_DIR}")
        return

    # Parsear y validar nombres
    tareas = []
    for archivo in archivos:
        resultado = parsear_nombre_archivo(archivo.name)
        if resultado is None:
            print(f"\n  AVISO: No se pudo parsear '{archivo.name}'")
            print("    Formato esperado: REPORTE-MES-SUCURSAL-AÑO.xlsx")
            continue
        mes_nombre, mes_num, sucursal, anio = resultado
        tareas.append((archivo.name, sucursal, mes_nombre, mes_num, anio))

    if not tareas:
        print("\nNo se encontraron archivos con formato válido.")
        return

    # Mostrar archivos detectados
    print(f"\nArchivos detectados ({len(tareas)}):")
    for nombre, sucursal, mes_nombre, _, anio in tareas:
        print(f"  - {nombre}  →  {sucursal} / {mes_nombre} {anio}")

    # Procesar
    for nombre, sucursal, mes_nombre, mes_num, anio in tareas:
        procesar_archivo(nombre, sucursal, mes_nombre, mes_num, anio)

    print("\n" + "=" * 60)
    print("PROCESO COMPLETADO")
    print("=" * 60)


if __name__ == "__main__":
    main()
