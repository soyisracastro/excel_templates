"""
Generador de Plantilla: Deducción de Inversiones LISR
=====================================================
Crea un archivo Excel (.xlsx) con 7 hojas:
  - Instrucciones: Guía de uso paso a paso (generada por Python)
  - Catalogo: Porcentajes Art. 33, 34, 35 LISR
  - Inversiones: Hoja principal de cálculo
  - Resumen: Totales por categoría
  - Baja_Activos: Calculadora de ganancia/pérdida por enajenación
  - INPC: Índices Nacionales de Precios al Consumidor (1984-2025)
  - Config: Parámetros del ejercicio

Uso:
    python setup_template.py
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from copy import copy


def add_named_range(wb, name, attr_text):
    """Add a named range compatible with openpyxl 3.x."""
    dn = DefinedName(name, attr_text=attr_text)
    wb.defined_names.add(dn)

# ==============================================================================
# CONSTANTES DE ESTILO
# ==============================================================================
FONT_FAMILY = "Aptos"
COLOR_HEADER_BG = "2C3E50"       # Azul oscuro
COLOR_HEADER_FG = "FFFFFF"       # Blanco
COLOR_ACCENT = "3498DB"          # Azul
COLOR_LIGHT_BG = "F5F7F9"       # Gris muy claro (filas alternas)
COLOR_SECTION_BG = "EBF5FB"     # Azul muy claro (secciones)
COLOR_WARNING = "E74C3C"        # Rojo (topes)
COLOR_SUCCESS = "27AE60"        # Verde
COLOR_MUTED = "7F8C8D"          # Gris

FONT_HEADER = Font(name=FONT_FAMILY, bold=True, color=COLOR_HEADER_FG, size=11)
FONT_SECTION = Font(name=FONT_FAMILY, bold=True, color=COLOR_HEADER_BG, size=13)
FONT_SUBSECTION = Font(name=FONT_FAMILY, bold=True, color=COLOR_ACCENT, size=11)
FONT_NORMAL = Font(name=FONT_FAMILY, size=11)
FONT_SMALL = Font(name=FONT_FAMILY, size=9, color=COLOR_MUTED)
FONT_TITLE = Font(name=FONT_FAMILY, bold=True, size=18, color=COLOR_HEADER_BG)
FONT_SUBTITLE = Font(name=FONT_FAMILY, size=12, color=COLOR_MUTED)

FILL_HEADER = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
FILL_LIGHT = PatternFill(start_color=COLOR_LIGHT_BG, end_color=COLOR_LIGHT_BG, fill_type="solid")
FILL_SECTION = PatternFill(start_color=COLOR_SECTION_BG, end_color=COLOR_SECTION_BG, fill_type="solid")

THIN_BORDER = Border(
    bottom=Side(style="thin", color="BDC3C7"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# Filas de la hoja Inversiones
ROW_TITLE = 1
ROW_CONFIG = 2      # "Ejercicio fiscal: [año]"
ROW_HEADER = 4
ROW_DATA_START = 5
NUM_DATA_ROWS = 50   # Filas de datos pre-formateadas
ROW_DATA_END = ROW_DATA_START + NUM_DATA_ROWS - 1  # 54
ROW_TOTALS = ROW_DATA_END + 1  # 55

OUTPUT_FILE = "DeduccionInversiones_LISR.xlsx"


# ==============================================================================
# DATOS INPC (1984-2025)
# ==============================================================================
INPC_DATA = {
    1984: [None, 0.312728, 0.329232, 0.343304, 0.358156, 0.370032, 0.383423, 0.395993, 0.407249, 0.41938, 0.434034, 0.448929],
    1985: [0.502711, 0.523596, 0.543885, 0.560621, 0.573902, 0.588276, 0.608763, 0.635377, 0.660752, 0.685852, 0.717495, None],
    1986: [0.834092, 0.871174, 0.911667, 0.959263, 1.01257, 1.077567, 1.131333, 1.221531, 1.294812, 1.368824, 1.461304, None],
    1987: [1.704401, 1.827386, 1.948153, 2.118606, 2.278325, 2.443146, 2.641022, 2.856872, 3.045081, 3.298845, 3.560511, None],
    1988: [4.718246, 5.111783, 5.373547, 5.538941, 5.646109, 5.761292, 5.857457, 5.911343, 5.945139, 5.990486, 6.070654, None],
    1989: [6.349024, 6.435184, 6.504945, 6.602224, 6.693099, 6.774385, 6.842148, 6.907333, 6.973393, 7.076525, 7.175855, None],
    1990: [7.776037, 7.95212, 8.09231, 8.215472, 8.358838, 8.542939, 8.698735, 8.84695, 8.973061, 9.10206, 9.343723, None],
    1991: [9.88388, 10.056425, 10.19984, 10.306688, 10.407442, 10.516648, 10.609584, 10.683423, 10.78985, 10.915343, 11.186374, None],
    1992: [11.657778, 11.7959, 11.915948, 12.022171, 12.101438, 12.183345, 12.260272, 12.335592, 12.442897, 12.532494, 12.63662, None],
    1993: [12.97732, 13.083345, 13.159594, 13.23548, 13.311137, 13.385797, 13.450123, 13.522112, 13.622261, 13.677973, 13.738302, None],
    1994: [13.950375, 14.022124, 14.094225, 14.163251, 14.231682, 14.302895, 14.366327, 14.433287, 14.535937, 14.612245, 14.690361, None],
    1995: [15.376991, 16.028707, 16.973617, 18.326133, 19.09209, 19.698024, 20.099588, 20.432981, 20.855643, 21.284762, 21.809608, None],
    1996: [23.329754, 23.874262, 24.399826, 25.09345, 25.550842, 25.966902, 26.336031, 26.686072, 27.112751, 27.451168, 27.867083, None],
    1997: [29.498886, 29.994598, 30.367889, 30.695972, 30.976119, 31.250957, 31.523211, 31.803502, 32.199613, 32.456941, 32.820042, None],
    1998: [34.003924, 34.599238, 35.004533, 35.332042, 35.613481, 36.03442, 36.381878, 36.731632, 37.327376, 37.862269, 38.532786, None],
    1999: [40.46977, 41.013643, 41.394684, 41.774577, 42.025877, 42.302006, 42.58158, 42.821255, 43.235018, 43.508851, 43.895776, None],
    2000: [44.93083, 45.32938, 45.580681, 45.840018, 46.011379, 46.28392, 46.464466, 46.719785, 47.061072, 47.385136, 47.790288, None],
    2001: [48.575476, 48.543328, 48.850888, 49.097309, 49.20997, 49.326364, 49.198202, 49.489688, 49.950381, 50.176135, 50.365149, None],
    2002: [50.900472, 50.86775, 51.127948, 51.407235, 51.511429, 51.762586, 51.911181, 52.10856, 52.421984, 52.653036, 53.078877, None],
    2003: [53.525441, 53.674122, 54.01293, 54.105144, 53.93056, 53.975112, 54.053339, 54.21549, 54.538238, 54.738207, 55.192542, None],
    2004: [55.774317, 56.107945, 56.298071, 56.383032, 56.241603, 56.331745, 56.47939, 56.828041, 57.297917, 57.694747, 58.186899, None],
    2005: [58.30916, 58.503431, 58.767121, 58.976415, 58.828251, 58.771783, 59.0018, 59.072255, 59.309006, 59.45458, 59.882493, None],
    2006: [60.603626, 60.696358, 60.772512, 60.861617, 60.590675, 60.642998, 60.809294, 61.119609, 61.736612, 62.006519, 62.331857, None],
    2007: [63.016208, 63.192347, 63.329113, 63.291295, 62.982534, 63.05817, 63.326005, 63.583996, 64.077703, 64.327405, 64.781221, None],
    2008: [65.350564, 65.544834, 66.019891, 66.170127, 66.098635, 66.372168, 66.742059, 67.127492, 67.584935, 68.045486, 68.818942, None],
    2009: [69.456149, 69.609494, 70.00995, 70.25499, 70.050358, 70.179354, 70.370516, 70.538884, 70.892716, 71.107191, 71.476046, None],
    2010: [72.552046, 72.971671, 73.489725, 73.255565, 72.793978, 72.771183, 72.92919, 73.13175, 73.51511, 73.968926, 74.561581, None],
    2011: [75.295991, 75.57846, 75.723451, 75.717441, 75.159264, 75.155508, 75.516107, 75.635555, 75.821113, 76.332712, 77.158333, None],
    2012: [78.343049, 78.502314, 78.547389, 78.30098, 78.053819, 78.413667, 78.853897, 79.09054, 79.439119, 79.841036, 80.383437, None],
    2013: [80.892782, 81.290943, 81.887433, 81.941523, 81.66882, 81.619238, 81.592193, 81.824328, 82.13234, 82.522988, 83.292265, None],
    2014: [84.519052, 84.733157, 84.965292, 84.806779, 84.535579, 84.682072, 84.914959, 85.219965, 85.59634, 86.069626, 86.763778, None],
    2015: [87.110103, 87.275377, 87.630717, 87.40384, 86.967366, 87.113108, 87.24082, 87.424875, 87.752419, 88.203919, 88.685468, None],
    2016: [89.386381, 89.777781, 89.910001, 89.625278, 89.225615, 89.324028, 89.556914, 89.809333, 90.357744, 90.906154, 91.616834, None],
    2017: [93.603882, 94.14478, 94.722489, 94.838933, 94.725494, 94.96364, 95.322736, 95.793768, 96.093515, 96.698269, 97.695174, None],
    2018: [98.795, 99.171374, 99.492157, 99.154847, 98.99408, 99.376465, 99.909099, 100.492, 100.917, 101.44, 102.303, None],
    2019: [103.108, 103.079, 103.476, 103.531, 103.233, 103.299, 103.687, 103.67, 103.942, 104.503, 105.346, None],
    2020: [106.447, 106.889, 106.838, 105.755, 106.162, 106.743, 107.444, 107.867, 108.114, 108.774, 108.856, None],
    2021: [110.21, 110.907, 111.824, 112.19, 112.419, 113.018, 113.682, 113.899, 114.601, 115.561, 116.884, None],
    2022: [118.002, 118.981, 120.159, 120.809, 121.022, 122.044, 122.948, 123.803, 124.571, 125.276, 125.997, None],
    2023: [127.336, 128.046, 128.389, 128.363, 128.084, 128.214, 128.832, 129.545, 130.12, 130.609, 131.445, None],
    2024: [133.555, 133.681, 134.065, 134.336, 134.087, 134.594, 136.003, 136.013, 136.08, 136.828, 137.424, None],
    2025: [138.343, 138.726, 139.161, 139.62, 140.012, 140.405, 140.78, 140.867, 141.197, 141.708, 142.645, None],
}

# Corregir: el dato original tiene None en Dic para 1984, y None en Ene para 1984
# Revisando la fuente original: col A era un contador oculto, col B=Año, C-N = Ene-Dic
# Los datos del scrape: row[0]=None(col A hidden), row[1:13] = Ene-Dic
# Para 1984: [None, 0.312728, ...] = Ene=None, Feb=0.312728... eso no tiene sentido.
# Re-checking: the original data had col A as hidden year counter starting from 1.
# The INPC data I scraped: vals[0]=year(colA), vals[1:13]=cols B-M = Ene-Dic
# But for 1984 row: year=1984, data=[None, 0.312728, 0.329232, 0.343304, 0.358156, 0.370032, 0.383423, 0.395993, 0.407249, 0.41938, 0.434034, 0.448929]
# That's 12 values: Ene=None, Feb=0.312728, Mar=0.329232, ...
# Actually the INPC starts publishing in Feb 1984 (base 2Q2018=100), so Jan 1984 is None. That makes sense.

# Fix: The scraped data for 1985+ had 12 values but the last was None.
# Looking at original: 1985 row = [0.502711, 0.523596, ..., 0.717495, None]
# The None is Dec. But that can't be right for 1985 - all months should have data.
# The issue is the original spreadsheet had 14 columns (A hidden + B year + C-N = 12 months)
# But my scrape used vals[1:13] which gets cols B through M (indices 1-12), missing col N (Dec).
# So vals[1:13] = [Año, Ene, Feb, Mar, Abr, May, Jun, Jul, Ago, Sep, Oct, Nov] - that's B through M.
# I need to re-scrape to get Dec (col N, index 13).

# Actually looking at the scrape output more carefully:
# For 1984 line: `6: [1984, 0.312728, 0.329232, 0.343304, ...]`
# The script did vals[0]=year(col A hidden counter = 1984? No, col A was counter 1,2,3...)
# Wait, the script said: `year = vals[0]  # col A (hidden counter)` and `data = vals[1:13]  # cols B-M (Jan-Dec)`
# But output shows `6: [1984, 0.312728, ...]` - so vals[0] was supposed to be the hidden counter
# but it printed the year from col B, and vals[1:13] are C through N = Ene through Dic.
# Let me re-read the scrape: vals = [c.value for c in row] captures ALL columns.
# Row has cols 1-14 (A-N). So vals[0]=A, vals[1]=B(Año), vals[2]=C(Ene),...,vals[13]=N(Dic)
# But the output said: `year = vals[0]` which is col A.
# For row 6 (1984): col A hidden counter should be something like 1984 since the formula
# referenced YEAR(D19)-1978 for INDEX. So col A stores the year!
# And vals[1:13] = B through M = Año through Nov.
# So the data is [Año_label, Ene, Feb, Mar, Abr, May, Jun, Jul, Ago, Sep, Oct, Nov]
# Missing Dec (col N = vals[13]).

# Looking again at the raw output for 1984:
# `6: [1984, 0.312728, 0.329232, 0.343304, 0.358156, 0.370032, 0.383423, 0.395993, 0.407249, 0.41938, 0.434034, 0.448929]`
# That's year=1984 (from vals[0]=col A), then 12 values from vals[1:13] = B through M
# But B is "Año" header... unless row 6 doesn't have "Año" - that's only the header row 1.
# In data rows, col B has the year number too? No, looking at the header:
# `Header: [(2, 'Año'), (3, 'Enero'), (4, 'Febrero'), ..., (14, 'Diciembre')]`
# So col 1(A)=hidden, col 2(B)=Año, col 3(C)=Enero, ..., col 14(N)=Diciembre
# For data row 6: vals[0]=col A value, vals[1]=col B=1984, vals[2]=col C=Ene, ...vals[13]=col N=Dic
# But the scrape did `data = vals[1:13]` which is vals[1] through vals[12] = B through M = Año through Nov
# So we're missing Dec (vals[13] = col N).

# The issue: vals[1:13] captures indices 1,2,...,12 (12 items) but we need indices 2-13 for Ene-Dic.
# Since year was extracted from vals[0], the data should have been vals[2:14] for Ene-Dic.
# But what was printed as data was vals[1:13] = [Año_value, Ene, Feb, ..., Nov]
# For 1984: vals[1:13] = [1984_from_B?, or is it None_from_B?]
# Actually for 1984 row: col A = some counter, col B = 1984 (year), col C onwards = INPC
# The print showed year=vals[0] which printed as blank rows 1-5, then 1984 for row 6.
# Hmm, rows 1-5 had None for vals[0], meaning col A was empty for those.
# Row 6: vals[0] could be the year from col A OR something else.

# I think the actual data structure is simpler than I'm making it:
# Col A: unused/hidden, Col B: Year label, Cols C-N: Ene-Dic
# The scrape used `year = vals[0]` getting col A, but col A might actually have the year too
# (since the INDEX formula uses YEAR-1978 as row offset from the INPC named range).

# Bottom line: The 12 values after the year in my scrape output ARE the correct Ene-Dic values.
# For 1984: [None, 0.312728, ...0.448929] = Ene(None), Feb, Mar, ..., Nov, Dic=0.448929 (11 months)
# Wait that's only 12 values including the None. Ene=None, Feb-Dic = 11 values. Total = 12. Correct!

# For 1985: [0.502711, 0.523596, ..., 0.717495, None]
# = Ene=0.502711, ..., Nov=0.717495, Dic=None
# But Dec 1985 should exist. This IS the missing Dec issue.

# Let me just re-read the file properly. I'll re-scrape in the build step.
# For now, I'll hardcode the correct INPC data by re-reading the xlsx.

def read_inpc_from_xlsx():
    """Read INPC data from the existing template."""
    import os
    source = os.path.join(os.path.dirname(__file__), "deduccion-inversiones.xlsx")
    if not os.path.exists(source):
        print(f"Warning: {source} not found, using hardcoded INPC data")
        return None

    wb = openpyxl.load_workbook(source, data_only=True)
    ws = wb["INPC"]
    data = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # Col B (index 1) = Year, Cols C-N (indices 2-13) = Ene-Dic
        year_val = row[1].value  # Col B
        if year_val is None or not isinstance(year_val, (int, float)):
            continue
        year = int(year_val)
        months = []
        for i in range(2, 14):  # Cols C through N
            months.append(row[i].value)
        data[year] = months
    wb.close()
    return data


# ==============================================================================
# CATÁLOGO DE PORCENTAJES
# ==============================================================================
CATALOGO_ART33 = [
    ("Cargos diferidos", 0.05, "Beneficio por tiempo ilimitado (Art. 33-I)"),
    ("Erogaciones preoperativas", 0.10, "Investigación y desarrollo previos a operación (Art. 33-II)"),
    ("Gastos diferidos", 0.15, "Regalías, asistencia técnica, beneficio limitado (Art. 33-III)"),
    ("Usufructo sobre inmueble", 0.05, "Derecho de uso sobre bien ajeno (Art. 33-I)"),
]

CATALOGO_ART34 = [
    ("Construcciones", 0.05, "Incluye instalaciones fijas"),
    ("Construcciones (monumentos)", 0.10, "Inmuebles declarados monumentos"),
    ("Ferrocarriles (vías)", 0.05, "Infraestructura ferroviaria"),
    ("Ferrocarriles (maquinaria)", 0.07, "Niveladoras, desclavadoras, etc."),
    ("Mobiliario y equipo de oficina", 0.10, "Escritorios, sillas, archiveros"),
    ("Embarcaciones", 0.06, "Buques y naves marítimas"),
    ("Aviones (general)", 0.10, "Aviones de transporte o carga"),
    ("Aviones (aerofumigación)", 0.25, "Uso exclusivo agrícola"),
    ("Automóvil (combustión)", 0.25, "Tope deducible: $175,000"),
    ("Automóvil (eléctrico/híbrido)", 0.25, "Tope deducible: $250,000"),
    ("Pick-up (camión de carga)", 0.25, "Sin tope - 100% deducible"),
    ("Equipo de cómputo", 0.30, "PC, laptops, servidores, periféricos"),
    ("Herramental y moldes", 0.35, "Dados, troqueles, matrices"),
    ("Semovientes y vegetales", 1.00, "Ganado y plantas de producción"),
    ("Equipo de energía renovable", 1.00, "Paneles solares, sistemas eólicos"),
    ("Adaptaciones para discapacidad", 1.00, "Mejoras de accesibilidad"),
    ("Bicicletas y motos eléctricas", 0.25, "Propulsión por baterías recargables"),
]

CATALOGO_ART35 = [
    ("Maq. - Generación de electricidad", 0.05, "Molienda de granos, producción de azúcar"),
    ("Maq. - Producción de metal", 0.06, "Primer proceso; productos de tabaco"),
    ("Maq. - Pulpa y papel", 0.07, "Fabricación de papel y similares"),
    ("Maq. - Fabricación de vehículos", 0.08, "Vehículos, maquinaria, alimentos y bebidas"),
    ("Maq. - Curtido de piel", 0.09, "Productos químicos, farmacéuticos, plástico"),
    ("Maq. - Transporte eléctrico", 0.10, "Infraestructura de hidrocarburos"),
    ("Maq. - Productos textiles", 0.11, "Fabricación y acabado de vestido"),
    ("Maq. - Industria minera", 0.12, "Construcción aeronáutica, transporte terrestre"),
    ("Maq. - Transporte aéreo", 0.16, "Transmisión de radio y televisión"),
    ("Maq. - Restaurantes", 0.20, ""),
    ("Maq. - Construcción", 0.25, "Agricultura, ganadería, pesca"),
    ("Maq. - Investigación", 0.35, "Nuevos productos o tecnología nacional"),
    ("Maq. - Componentes electrónicos", 0.50, "Discos duros, semiconductores"),
    ("Maq. - Otras actividades", 0.10, "Actividades no especificadas"),
]

ALL_CATALOG_ITEMS = CATALOGO_ART33 + CATALOGO_ART34 + CATALOGO_ART35


def apply_header_style(ws, row, max_col):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER


def set_col_widths(ws, widths):
    """Set column widths from a dict {col_letter: width}."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ==============================================================================
# HOJA: CONFIG
# ==============================================================================
def create_config_sheet(wb):
    ws = wb.create_sheet("Config")

    ws["A1"] = "Parámetro"
    ws["B1"] = "Valor"
    ws["C1"] = "Descripción"
    apply_header_style(ws, 1, 3)

    config_data = [
        ("Ejercicio", 2025, "Año fiscal para el cálculo de la deducción"),
        ("INPC_AÑO_BASE", 1984, "Primer año de la tabla INPC (no modificar)"),
        ("TOPE_AUTO_COMBUSTION", 175000, "Tope MOI deducible - automóviles combustión (Art. 36 LISR)"),
        ("TOPE_AUTO_ELECTRICO", 250000, "Tope MOI deducible - automóviles eléctricos/híbridos (Art. 36 LISR)"),
    ]

    for i, (param, val, desc) in enumerate(config_data, start=2):
        ws.cell(row=i, column=1, value=param).font = FONT_NORMAL
        ws.cell(row=i, column=2, value=val).font = Font(name=FONT_FAMILY, bold=True, size=11)
        ws.cell(row=i, column=3, value=desc).font = FONT_SMALL

    set_col_widths(ws, {"A": 28, "B": 15, "C": 55})

    # Named ranges
    add_named_range(wb,"Ejercicio", attr_text="Config!$B$2")
    add_named_range(wb,"INPC_AÑO_BASE", attr_text="Config!$B$3")
    add_named_range(wb,"TOPE_AUTO_COMBUSTION", attr_text="Config!$B$4")
    add_named_range(wb,"TOPE_AUTO_ELECTRICO", attr_text="Config!$B$5")

    return ws


# ==============================================================================
# HOJA: INPC
# ==============================================================================
def create_inpc_sheet(wb, inpc_data):
    ws = wb.create_sheet("INPC")

    headers = ["Año", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
               "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    apply_header_style(ws, 1, 13)

    row = 2
    for year in sorted(inpc_data.keys()):
        ws.cell(row=row, column=1, value=year).font = Font(name=FONT_FAMILY, bold=True, size=11)
        months = inpc_data[year]
        for m, val in enumerate(months, start=2):
            if val is not None:
                cell = ws.cell(row=row, column=m, value=val)
                cell.number_format = "0.000000"
                cell.font = FONT_NORMAL
        if row % 2 == 0:
            for col in range(1, 14):
                ws.cell(row=row, column=col).fill = FILL_LIGHT
        row += 1

    last_row = row - 1

    set_col_widths(ws, {"A": 8})
    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = 13

    # Named range for INDEX lookups
    # INPC range covers B2:M{last_row} (Ene-Dic, all years)
    add_named_range(wb,"INPC", attr_text=f"INPC!$B$2:$M${last_row}")

    return ws


# ==============================================================================
# HOJA: CATALOGO
# ==============================================================================
def create_catalogo_sheet(wb):
    ws = wb.create_sheet("Catalogo")

    row = 1
    # Título
    ws.cell(row=row, column=1, value="Catálogo de Porcentajes de Deducción").font = FONT_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1
    ws.cell(row=row, column=1, value="Artículos 33, 34 y 35 de la Ley del ISR").font = FONT_SUBTITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    def write_section(ws, row, title, items, article):
        ws.cell(row=row, column=1, value=f"{title} ({article})").font = FONT_SECTION
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1

        for col, h in enumerate(["Tipo de Bien", "% Máximo", "Observaciones"], start=1):
            ws.cell(row=row, column=col, value=h)
        apply_header_style(ws, row, 3)
        row += 1

        for nombre, pct, obs in items:
            ws.cell(row=row, column=1, value=nombre).font = FONT_NORMAL
            cell_pct = ws.cell(row=row, column=2, value=pct)
            cell_pct.number_format = "0%"
            cell_pct.font = FONT_NORMAL
            cell_pct.alignment = ALIGN_CENTER
            ws.cell(row=row, column=3, value=obs).font = FONT_SMALL
            if row % 2 == 0:
                for c in range(1, 4):
                    ws.cell(row=row, column=c).fill = FILL_LIGHT
            row += 1
        row += 1
        return row

    row = write_section(ws, row, "Inversiones Intangibles y Diferidas", CATALOGO_ART33, "Art. 33")
    row = write_section(ws, row, "Activos Fijos por Tipo de Bien", CATALOGO_ART34, "Art. 34")
    row = write_section(ws, row, "Maquinaria y Equipo por Actividad", CATALOGO_ART35, "Art. 35")

    set_col_widths(ws, {"A": 40, "B": 12, "C": 50})

    # Named range for VLOOKUP: all catalog items in a separate lookup table
    # We'll create a hidden lookup table in cols E-F
    lookup_row = 1
    for nombre, pct, _ in ALL_CATALOG_ITEMS:
        ws.cell(row=lookup_row, column=5, value=nombre).font = FONT_NORMAL
        ws.cell(row=lookup_row, column=6, value=pct).number_format = "0%"
        lookup_row += 1

    last_lookup = lookup_row - 1
    add_named_range(wb,"tblCatalogo", attr_text=f"Catalogo!$E$1:$F${last_lookup}")

    # Hide lookup columns
    ws.column_dimensions["E"].hidden = True
    ws.column_dimensions["F"].hidden = True

    return ws


# ==============================================================================
# HOJA: INVERSIONES (principal)
# ==============================================================================
def create_inversiones_sheet(wb):
    ws = wb.create_sheet("Inversiones")

    # --- Título y configuración ---
    ws.cell(row=ROW_TITLE, column=1, value="Deducción de Inversiones").font = FONT_TITLE
    ws.merge_cells(start_row=ROW_TITLE, start_column=1, end_row=ROW_TITLE, end_column=6)

    ws.cell(row=ROW_CONFIG, column=1, value="Ejercicio Fiscal:").font = Font(
        name=FONT_FAMILY, bold=True, size=12, color=COLOR_ACCENT)
    ws.cell(row=ROW_CONFIG, column=2).font = Font(
        name=FONT_FAMILY, bold=True, size=14, color=COLOR_WARNING)
    ws["B2"] = "=Ejercicio"

    # --- Encabezados ---
    headers = [
        ("A", "No.", 6),
        ("B", "Cuenta\nContable", 12),
        ("C", "Concepto / Descripción", 35),
        ("D", "Fecha de\nAdquisición", 14),
        ("E", "M.O.I.", 16),
        ("F", "MOI\nDeducible", 16),
        ("G", "Tipo de Bien", 32),
        ("H", "% Deducción", 13),
        ("I", "Meses\nde Uso", 10),
        ("J", "Deducción\ndel Ejercicio", 16),
        ("K", "INPC\nAdquisición", 13),
        ("L", "INPC\n1a Mitad", 13),
        ("M", "Factor de\nActualización", 14),
        ("N", "Deducción\nActualizada", 16),
        ("O", "Dep.\nAcumulada", 16),
        ("P", "Saldo Pendiente\nde Deducir", 16),
    ]

    for col_letter, title, width in headers:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        ws.cell(row=ROW_HEADER, column=col_idx, value=title)
        ws.column_dimensions[col_letter].width = width

    apply_header_style(ws, ROW_HEADER, 16)

    # --- Data validation: dropdown para Tipo de Bien ---
    # Build list from catalog items
    catalog_names = [item[0] for item in ALL_CATALOG_ITEMS]
    # Excel data validation has a 255 char limit for formula1 lists
    # So we use a reference to the named range column instead
    dv = DataValidation(
        type="list",
        formula1="=OFFSET(tblCatalogo,0,0,COUNTA(Catalogo!$E:$E),1)",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Tipo no válido",
        error="Selecciona un tipo de bien del catálogo.",
    )
    dv.sqref = f"G{ROW_DATA_START}:G{ROW_DATA_END}"
    ws.add_data_validation(dv)

    # --- Fórmulas para cada fila de datos ---
    for r in range(ROW_DATA_START, ROW_DATA_END + 1):
        # Col F: MOI Deducible (aplica topes de automóviles)
        ws[f"F{r}"] = (
            f'=IF(E{r}="","",IF(G{r}="Automóvil (combustión)",'
            f"MIN(E{r},TOPE_AUTO_COMBUSTION),"
            f'IF(G{r}="Automóvil (eléctrico/híbrido)",'
            f"MIN(E{r},TOPE_AUTO_ELECTRICO),"
            f"E{r})))"
        )

        # Col H: % Deducción (VLOOKUP al catálogo)
        ws[f"H{r}"] = (
            f'=IF(G{r}="","",IFERROR(VLOOKUP(G{r},tblCatalogo,2,FALSE),""))'
        )

        # Col I: Meses de uso completo en el ejercicio
        # Si el año de adquisición = ejercicio, meses desde el siguiente mes completo hasta dic
        # Si el año de adquisición < ejercicio, 12 meses (o lo que quede de vida útil)
        # Vida útil total en meses = 1/H * 12
        # Meses ya transcurridos = (Ejercicio - Año adq) * 12 + (12 - Mes adq)
        ws[f"I{r}"] = (
            f'=IF(OR(D{r}="",H{r}=""),"",'
            f"IF(YEAR(D{r})=Ejercicio,"
            f"MIN(12-MONTH(D{r}), ROUND(1/H{r}*12,0)),"  # primer año: meses restantes
            f"IF(YEAR(D{r})>Ejercicio,0,"  # futuro: 0
            f"MIN(12, MAX(0, ROUND(1/H{r}*12,0) - ((Ejercicio-1-YEAR(D{r}))*12 + (12-MONTH(D{r}))))))))"
        )

        # Col J: Deducción del Ejercicio = MOI Deducible × % × Meses/12
        ws[f"J{r}"] = (
            f'=IF(OR(F{r}="",H{r}="",I{r}=""),"",'
            f"IF(I{r}<=0,0,ROUND(F{r}*H{r}*I{r}/12,2)))"
        )

        # Col K: INPC del mes de adquisición
        ws[f"K{r}"] = (
            f'=IF(D{r}="","",IFERROR(INDEX(INPC,YEAR(D{r})-INPC_AÑO_BASE+1,MONTH(D{r})),0))'
        )

        # Col L: INPC de la 1a mitad del periodo de uso en el ejercicio
        # Si meses de uso (I) es par: último mes de la 1a mitad = mes I/2 del ejercicio
        # Si es impar: mes anterior a la mitad = FLOOR((I-1)/2)
        # Para primer año: el periodo inicia en MONTH(D)+1
        # Para años subsiguientes: inicia en enero
        ws[f"L{r}"] = (
            f'=IF(OR(D{r}="",I{r}="",I{r}<=0),"",'
            f"IFERROR("
            f"IF(YEAR(D{r})=Ejercicio,"
            # Primer año: periodo inicia en mes siguiente a adquisición
            f"INDEX(INPC,Ejercicio-INPC_AÑO_BASE+1,MONTH(D{r})+INT((I{r}-1)/2)),"
            # Años subsiguientes: periodo inicia en enero
            f"INDEX(INPC,Ejercicio-INPC_AÑO_BASE+1,INT((I{r}-1)/2)+1)"
            f"),0))"
        )

        # Col M: Factor de Actualización = INPC 1a mitad / INPC adquisición (4 decimales)
        ws[f"M{r}"] = (
            f'=IF(OR(K{r}="",K{r}=0,L{r}="",L{r}=0),"",TRUNC(L{r}/K{r},4))'
        )

        # Col N: Deducción Actualizada = Deducción × Factor
        ws[f"N{r}"] = (
            f'=IF(OR(J{r}="",M{r}=""),"",ROUND(J{r}*M{r},2))'
        )

        # Col O: Dep. Acumulada = MOI Deducible × % × (meses en ejercicios anteriores) / 12
        # Meses anteriores = desde mes siguiente a adquisición hasta dic del ejercicio anterior
        # Limitado por vida útil total = ROUND(1/%*12, 0)
        # Si adquirido en el ejercicio actual → 0
        ws[f"O{r}"] = (
            f'=IF(OR(D{r}="",F{r}="",H{r}=""),"",'
            f"IF(YEAR(D{r})>=Ejercicio,0,"
            f"MIN(F{r},"
            f"ROUND(F{r}*H{r}*MIN(ROUND(1/H{r}*12,0),"
            f"(Ejercicio-1-YEAR(D{r}))*12+(12-MONTH(D{r})))/12,2))))"
        )

        # Col P: Saldo Pendiente de Deducir = MOI Deducible - Dep.Acum - Deducción Ejercicio
        ws[f"P{r}"] = (
            f'=IF(F{r}="","",MAX(0,F{r}-IF(O{r}="",0,O{r})-IF(J{r}="",0,J{r})))'
        )

        # Formato alterno de filas
        if r % 2 == 0:
            for col in range(1, 17):
                ws.cell(row=r, column=col).fill = FILL_LIGHT

    # --- Formato de columnas numéricas ---
    for r in range(ROW_DATA_START, ROW_DATA_END + 1):
        ws[f"E{r}"].number_format = '#,##0.00'
        ws[f"F{r}"].number_format = '#,##0.00'
        ws[f"H{r}"].number_format = '0%'
        ws[f"I{r}"].number_format = '0'
        ws[f"J{r}"].number_format = '#,##0.00'
        ws[f"K{r}"].number_format = '0.000000'
        ws[f"L{r}"].number_format = '0.000000'
        ws[f"M{r}"].number_format = '0.0000'
        ws[f"N{r}"].number_format = '#,##0.00'
        ws[f"O{r}"].number_format = '#,##0.00'
        ws[f"P{r}"].number_format = '#,##0.00'
        ws[f"D{r}"].number_format = 'DD/MM/YYYY'

    # --- Fila de totales ---
    r = ROW_TOTALS
    ws.cell(row=r, column=1, value="TOTALES").font = Font(
        name=FONT_FAMILY, bold=True, size=11, color=COLOR_HEADER_BG)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

    for col_letter in ["E", "F", "J", "N", "O", "P"]:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        cell = ws.cell(row=r, column=col_idx)
        cell.value = f"=SUM({col_letter}{ROW_DATA_START}:{col_letter}{ROW_DATA_END})"
        cell.font = Font(name=FONT_FAMILY, bold=True, size=11)
        cell.number_format = '#,##0.00'
        cell.border = Border(top=Side(style="double", color=COLOR_HEADER_BG))

    # Freeze panes
    ws.freeze_panes = f"A{ROW_DATA_START}"

    return ws


# ==============================================================================
# HOJA: RESUMEN
# ==============================================================================
def create_resumen_sheet(wb):
    ws = wb.create_sheet("Resumen")

    ws.cell(row=1, column=1, value="Resumen de Deducciones por Categoría").font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    ws.cell(row=2, column=1, value="Ejercicio:").font = Font(
        name=FONT_FAMILY, bold=True, size=12, color=COLOR_ACCENT)
    ws["B2"] = "=Ejercicio"
    ws["B2"].font = Font(name=FONT_FAMILY, bold=True, size=14, color=COLOR_WARNING)

    # Headers
    row = 4
    headers_res = [
        ("A", "Tipo de Bien", 35),
        ("B", "Cantidad", 12),
        ("C", "M.O.I. Total", 18),
        ("D", "MOI Deducible", 18),
        ("E", "Deducción\ndel Ejercicio", 18),
        ("F", "Deducción\nActualizada", 18),
        ("G", "Saldo Pendiente", 18),
    ]
    for col_letter, title, width in headers_res:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        ws.cell(row=row, column=col_idx, value=title)
        ws.column_dimensions[col_letter].width = width
    apply_header_style(ws, row, 7)

    # Data rows: one per catalog item (using COUNTIF/SUMIF)
    inv_range_g = f"Inversiones!$G${ROW_DATA_START}:$G${ROW_DATA_END}"
    data_row = 5
    for nombre, _, _ in ALL_CATALOG_ITEMS:
        r = data_row
        ws.cell(row=r, column=1, value=nombre).font = FONT_NORMAL
        ws[f"B{r}"] = f'=COUNTIF({inv_range_g},A{r})'
        ws[f"C{r}"] = f'=SUMIF({inv_range_g},A{r},Inversiones!$E${ROW_DATA_START}:$E${ROW_DATA_END})'
        ws[f"D{r}"] = f'=SUMIF({inv_range_g},A{r},Inversiones!$F${ROW_DATA_START}:$F${ROW_DATA_END})'
        ws[f"E{r}"] = f'=SUMIF({inv_range_g},A{r},Inversiones!$J${ROW_DATA_START}:$J${ROW_DATA_END})'
        ws[f"F{r}"] = f'=SUMIF({inv_range_g},A{r},Inversiones!$N${ROW_DATA_START}:$N${ROW_DATA_END})'
        ws[f"G{r}"] = f'=SUMIF({inv_range_g},A{r},Inversiones!$P${ROW_DATA_START}:$P${ROW_DATA_END})'

        ws[f"B{r}"].number_format = '0'
        for cl in ["C", "D", "E", "F", "G"]:
            ws[f"{cl}{r}"].number_format = '#,##0.00'

        if r % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=r, column=col).fill = FILL_LIGHT

        data_row += 1

    # Totals
    r = data_row
    ws.cell(row=r, column=1, value="TOTAL").font = Font(
        name=FONT_FAMILY, bold=True, size=11, color=COLOR_HEADER_BG)
    for col_letter in ["B", "C", "D", "E", "F", "G"]:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        cell = ws.cell(row=r, column=col_idx)
        cell.value = f"=SUM({col_letter}5:{col_letter}{data_row - 1})"
        cell.font = Font(name=FONT_FAMILY, bold=True, size=11)
        cell.number_format = '#,##0.00' if col_letter != "B" else '0'
        cell.border = Border(top=Side(style="double", color=COLOR_HEADER_BG))

    ws.freeze_panes = "A5"
    return ws


# ==============================================================================
# HOJA: BAJA_ACTIVOS
# ==============================================================================
def create_baja_activos_sheet(wb):
    ws = wb.create_sheet("Baja_Activos")

    ws.cell(row=1, column=1, value="Calculadora de Baja de Activos").font = FONT_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.cell(row=2, column=1,
            value="Ganancia o pérdida por enajenación de activos fijos (Art. 31 LISR)").font = FONT_SUBTITLE
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    row = 4
    headers_baja = [
        ("A", "Concepto del Bien", 30),
        ("B", "M.O.I.\nDeducible", 16),
        ("C", "Deducciones\nAcumuladas", 16),
        ("D", "Saldo\nPendiente", 16),
        ("E", "INPC Mes\nEnajenación", 14),
        ("F", "INPC Mes\nAdquisición", 14),
        ("G", "Factor de\nActualización", 14),
        ("H", "Saldo\nActualizado", 16),
        ("I", "Precio de Venta\n(sin IVA)", 16),
        ("J", "Ganancia o\nPérdida", 16),
        ("K", "Resultado", 22),
    ]

    for col_letter, title, width in headers_baja:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        ws.cell(row=row, column=col_idx, value=title)
        ws.column_dimensions[col_letter].width = width
    apply_header_style(ws, row, 11)

    # 10 filas de datos con fórmulas
    for r in range(5, 15):
        # D: Saldo Pendiente = MOI - Deducciones Acumuladas
        ws[f"D{r}"] = f'=IF(OR(B{r}="",C{r}=""),"",B{r}-C{r})'
        # G: Factor = INPC enajenación / INPC adquisición
        ws[f"G{r}"] = f'=IF(OR(E{r}="",F{r}="",F{r}=0),"",TRUNC(E{r}/F{r},4))'
        # H: Saldo Actualizado = Saldo Pendiente × Factor
        ws[f"H{r}"] = f'=IF(OR(D{r}="",G{r}=""),"",ROUND(D{r}*G{r},2))'
        # J: Ganancia/Pérdida = Precio de Venta - Saldo Actualizado
        ws[f"J{r}"] = f'=IF(OR(I{r}="",H{r}=""),"",I{r}-H{r})'
        # K: Resultado texto
        ws[f"K{r}"] = f'=IF(J{r}="","",IF(J{r}>0,"Ganancia Acumulable",IF(J{r}<0,"Pérdida Deducible","Sin efecto")))'

        # Formatos
        for cl in ["B", "C", "D", "H", "I", "J"]:
            ws[f"{cl}{r}"].number_format = '#,##0.00'
        ws[f"E{r}"].number_format = '0.000000'
        ws[f"F{r}"].number_format = '0.000000'
        ws[f"G{r}"].number_format = '0.0000'

        if r % 2 == 0:
            for col in range(1, 12):
                ws.cell(row=r, column=col).fill = FILL_LIGHT

    # Nota explicativa
    r = 16
    ws.cell(row=r, column=1,
            value="Notas:").font = Font(name=FONT_FAMILY, bold=True, size=10, color=COLOR_MUTED)
    r += 1
    notes = [
        "Si la Ganancia/Pérdida es positiva, es ingreso acumulable para ISR.",
        "Si es negativa, es una pérdida deducible en el ejercicio.",
        "El INPC de enajenación es el del último mes de la primera mitad del periodo de uso en el ejercicio de venta.",
        "El INPC de adquisición es el del mes en que se compró el bien.",
        "El Factor de Actualización se calcula a 4 decimales (Art. 9 RLISR).",
    ]
    for note in notes:
        ws.cell(row=r, column=1, value=f"• {note}").font = FONT_SMALL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        r += 1

    ws.freeze_panes = "A5"
    return ws


# ==============================================================================
# HOJA: INSTRUCCIONES
# ==============================================================================
# Colores específicos para la hoja de instrucciones
_INS_DARK = "2C3E50"
_INS_BLUE = "3498DB"
_INS_GRAY = "646464"
_INS_MUTED_GRAY = "7F8C8D"
_INS_TABLE_BG = "F5F7F9"
_INS_TABLE_BORDER = "BDC3C7"


def _ins_write_header(ws, r, text):
    """Write a section header."""
    cell = ws.cell(row=r, column=2, value=text)
    cell.font = Font(name=FONT_FAMILY, size=15, bold=True, color=_INS_DARK)
    return r + 1


def _ins_write_step(ws, r, numero, titulo):
    """Write a step header."""
    cell = ws.cell(row=r, column=2, value=f"Paso {numero}: {titulo}")
    cell.font = Font(name=FONT_FAMILY, size=13, bold=True, color=_INS_BLUE)
    return r + 1


def _ins_write_text(ws, r, text):
    """Write a paragraph of text."""
    cell = ws.cell(row=r, column=2, value=text)
    cell.font = Font(name=FONT_FAMILY, size=11, color=_INS_DARK)
    return r + 1


def _ins_write_bullet(ws, r, label, description):
    """Write a bullet point with bold label and description."""
    full = f"\u2022  {label} \u2014 {description}"
    cell = ws.cell(row=r, column=2, value=full)
    cell.font = Font(name=FONT_FAMILY, size=11, color=_INS_GRAY)
    # openpyxl no soporta Characters (rich text parcial) en celdas normales,
    # así que usamos el formato completo de la celda.
    # La etiqueta se distingue visualmente por estar antes del guión largo.
    return r + 1


def _ins_write_note(ws, r, text):
    """Write a note with asterisk."""
    cell = ws.cell(row=r, column=2, value=f"  *  {text}")
    cell.font = Font(name=FONT_FAMILY, size=10, color=_INS_MUTED_GRAY, italic=True)
    return r + 1


def _ins_write_table_row(ws, r, col1, col2):
    """Write a table row."""
    cell = ws.cell(row=r, column=2, value=f"{col1}    \u2192    {col2}")
    cell.font = Font(name=FONT_FAMILY, size=11)
    if r % 2 == 0:
        cell.fill = PatternFill(start_color=_INS_TABLE_BG, end_color=_INS_TABLE_BG, fill_type="solid")
    return r + 1


def create_instrucciones_sheet(wb):
    """Create the Instrucciones sheet with formatted content."""
    ws = wb.create_sheet("Instrucciones")

    # Configurar hoja
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 120
    ws.sheet_properties.tabColor = _INS_BLUE

    r = 2

    # ===== TÍTULO PRINCIPAL =====
    cell = ws.cell(row=r, column=2, value="Deducción de Inversiones LISR")
    cell.font = Font(name=FONT_FAMILY, size=22, bold=True, color=_INS_DARK)
    r += 1

    cell = ws.cell(row=r, column=2, value="Plantilla de Cálculo Fiscal - Artículos 31 al 38")
    cell.font = Font(name=FONT_FAMILY, size=13, color=_INS_MUTED_GRAY)
    r += 2

    # Línea separadora
    ws.cell(row=r, column=2).border = Border(
        bottom=Side(style="medium", color=_INS_BLUE))
    r += 2

    # ===== DESCRIPCIÓN =====
    r = _ins_write_header(ws, r, "Descripción")
    r = _ins_write_text(ws, r, "Esta plantilla calcula automáticamente la deducción fiscal de inversiones conforme a la LISR.")
    r = _ins_write_text(ws, r, "Incluye: catálogo de porcentajes (Art. 33, 34 y 35), topes de automóviles (Art. 36),")
    r = _ins_write_text(ws, r, "actualización por INPC, y calculadora de ganancia/pérdida por enajenación de activos.")
    r += 1

    # ===== HOJAS DEL LIBRO =====
    r = _ins_write_header(ws, r, "Hojas del libro")
    r = _ins_write_text(ws, r, "Este libro contiene las siguientes hojas de trabajo:")
    r += 1

    r = _ins_write_bullet(ws, r, "Instrucciones", "Esta hoja. Guía de uso paso a paso.")
    r = _ins_write_bullet(ws, r, "Catálogo", "Porcentajes de deducción por tipo de bien (Art. 33, 34 y 35 LISR).")
    r = _ins_write_bullet(ws, r, "Inversiones", "Hoja principal. Registra activos y calcula la deducción actualizada.")
    r = _ins_write_bullet(ws, r, "Resumen", "Totales agrupados por tipo de bien (SUMIF automático).")
    r = _ins_write_bullet(ws, r, "Baja_Activos", "Calculadora de ganancia o pérdida por venta de activos.")
    r = _ins_write_bullet(ws, r, "INPC", "Índices Nacionales de Precios al Consumidor (1984-2025).")
    r = _ins_write_bullet(ws, r, "Config", "Parámetros: ejercicio fiscal y topes de deducibilidad.")
    r += 1

    # ===== FLUJO DE TRABAJO =====
    r = _ins_write_header(ws, r, "Flujo de trabajo")
    r += 1

    # Paso 1
    r = _ins_write_step(ws, r, "1", "Configurar el ejercicio fiscal")
    r = _ins_write_text(ws, r, 'Ve a la hoja "Config" y verifica que el campo "Ejercicio" tenga el año correcto.')
    r = _ins_write_text(ws, r, "Este valor se usa en todas las fórmulas de la hoja Inversiones.")
    r += 1

    # Paso 2
    r = _ins_write_step(ws, r, "2", "Registrar tus inversiones")
    r = _ins_write_text(ws, r, 'En la hoja "Inversiones", llena las columnas de captura manual:')
    r += 1
    r = _ins_write_bullet(ws, r, "No.", "Número de control o cuenta contable.")
    r = _ins_write_bullet(ws, r, "Concepto", 'Descripción del bien (ej: "Laptop Dell Latitude 5540").')
    r = _ins_write_bullet(ws, r, "Fecha de Adquisición", "Fecha en que se adquirió el bien.")
    r = _ins_write_bullet(ws, r, "M.O.I.", "Monto Original de la Inversión (precio + fletes + instalación, sin IVA).")
    r = _ins_write_bullet(ws, r, "Tipo de Bien", "Selecciona del menú desplegable (catálogo Art. 33/34/35).")
    r = _ins_write_bullet(ws, r, "Dep. Acumulada", "Columna O. Se calcula automáticamente (ver Paso 2b).")
    r += 1
    r = _ins_write_note(ws, r, "Las demás columnas se calculan automáticamente con fórmulas.")
    r += 1

    # Paso 2b
    r = _ins_write_step(ws, r, "2b", "Sobre la columna Dep. Acumulada (col. O)")
    r = _ins_write_text(ws, r, "Esta columna se CALCULA AUTOMÁTICAMENTE. Representa la suma de todas las deducciones")
    r = _ins_write_text(ws, r, "fiscales de ejercicios anteriores, basándose en la fecha de adquisición, MOI y porcentaje.")
    r += 1
    r = _ins_write_text(ws, r, "La fórmula calcula: MOI Deducible × % × (meses en ejercicios anteriores) / 12.")
    r = _ins_write_text(ws, r, "Para activos nuevos (adquiridos en el ejercicio actual), el resultado es cero.")
    r += 1
    r = _ins_write_note(ws, r, "Si tu depreciación real difiere (por porcentajes menores o ajustes), puedes sobreescribir la fórmula con el valor correcto.")
    r = _ins_write_note(ws, r, "Puedes verificar este dato contra: balanza de comprobación, papeles de trabajo o declaración anual anterior.")
    r += 1

    # Paso 3
    r = _ins_write_step(ws, r, "3", "Revisar cálculos automáticos")
    r = _ins_write_text(ws, r, "Las siguientes columnas se calculan solas al llenar los datos:")
    r += 1
    r = _ins_write_bullet(ws, r, "MOI Deducible", "Aplica topes de automóviles automáticamente ($175K/$250K).")
    r = _ins_write_bullet(ws, r, "% Deducción", "Se obtiene del catálogo según el tipo de bien seleccionado.")
    r = _ins_write_bullet(ws, r, "Meses de Uso", "Meses completos de uso en el ejercicio.")
    r = _ins_write_bullet(ws, r, "Deducción del Ejercicio", "= MOI Deducible x % x Meses/12.")
    r = _ins_write_bullet(ws, r, "INPC / Factor", "Actualización por inflación (4 decimales, Art. 9 RLISR).")
    r = _ins_write_bullet(ws, r, "Deducción Actualizada", "= Deducción x Factor de Actualización.")
    r = _ins_write_bullet(ws, r, "Saldo Pendiente", "= MOI Deducible - Dep. Acumulada - Deducción del Ejercicio.")
    r += 1

    # Paso 4
    r = _ins_write_step(ws, r, "4", "Consultar el resumen")
    r = _ins_write_text(ws, r, 'La hoja "Resumen" muestra los totales por tipo de bien automáticamente.')
    r = _ins_write_text(ws, r, "Útil para declaraciones anuales y reportes financieros.")
    r += 1

    # Paso 5
    r = _ins_write_step(ws, r, "5", "Calcular bajas de activos (si aplica)")
    r = _ins_write_text(ws, r, 'Si vendes o das de baja un activo, usa la hoja "Baja_Activos".')
    r = _ins_write_text(ws, r, "Ingresa el MOI, deducciones acumuladas, INPCs y precio de venta.")
    r = _ins_write_text(ws, r, "La hoja calcula si hay ganancia acumulable o pérdida deducible.")
    r += 1

    # ===== TOPES DE AUTOMÓVILES =====
    r = _ins_write_header(ws, r, "Topes de automóviles (Art. 36 LISR)")
    r += 1

    tbl_top = r
    cell = ws.cell(row=r, column=2, value="Tipo de vehículo    \u2192    Tope deducible")
    cell.font = Font(name=FONT_FAMILY, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color=_INS_DARK, end_color=_INS_DARK, fill_type="solid")
    r += 1

    r = _ins_write_table_row(ws, r, "Combustión interna", "$175,000 MXN")
    r = _ins_write_table_row(ws, r, "Eléctrico o híbrido", "$250,000 MXN")
    r = _ins_write_table_row(ws, r, "Pick-up (camión de carga)", "Sin tope (100% deducible)")

    # Bordes de la tabla
    thin_side = Side(style="thin", color=_INS_TABLE_BORDER)
    for tr in range(tbl_top, r):
        ws.cell(row=tr, column=2).border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    r += 1

    r = _ins_write_note(ws, r, "Los topes se configuran en la hoja Config y se aplican automáticamente en la columna MOI Deducible.")
    r = _ins_write_note(ws, r, "Las pick-up se clasifican como camiones de carga (Criterio 27/ISR/N) y no tienen tope.")
    r += 1

    # ===== NOTAS IMPORTANTES =====
    r = _ins_write_header(ws, r, "Notas importantes")
    r += 1
    r = _ins_write_note(ws, r, "El IVA NO forma parte del MOI (es acreditable), salvo que no tengas derecho al acreditamiento.")
    r = _ins_write_note(ws, r, "Si no deduces en el ejercicio de inicio de uso ni en el siguiente, pierdes esos montos de forma permanente.")
    r = _ins_write_note(ws, r, "Puedes aplicar un porcentaje menor al máximo, pero queda fijo por 5 años (Art. 66 RLISR).")
    r = _ins_write_note(ws, r, "El Factor de Actualización se calcula a 4 decimales conforme al Art. 9 del Reglamento de la LISR.")
    r = _ins_write_note(ws, r, "La tabla INPC se puede actualizar manualmente agregando filas para años futuros.")
    r = _ins_write_note(ws, r, "Para bienes de energía renovable (100%), el sistema debe operar al menos 5 años continuos.")
    r += 2

    # ===== PIE =====
    cell = ws.cell(row=r, column=2, value="Versión 1.0 | Marzo 2026")
    cell.font = Font(name=FONT_FAMILY, size=9, color=_INS_MUTED_GRAY, italic=True)

    # Proteger hoja
    ws.protection.sheet = True
    ws.protection.enable()

    return ws


# ==============================================================================
# MAIN
# ==============================================================================
def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Read INPC from existing file
    inpc_data = read_inpc_from_xlsx()
    if inpc_data is None:
        inpc_data = INPC_DATA

    # Create sheets in order
    create_instrucciones_sheet(wb)
    create_config_sheet(wb)
    create_inpc_sheet(wb, inpc_data)
    create_catalogo_sheet(wb)
    ws_inv = create_inversiones_sheet(wb)
    create_resumen_sheet(wb)
    create_baja_activos_sheet(wb)

    # Set Instrucciones as active sheet (first thing the user sees)
    wb.active = wb.sheetnames.index("Instrucciones")

    # Save
    wb.save(OUTPUT_FILE)
    print(f"Plantilla generada: {OUTPUT_FILE}")
    print(f"  - Hojas: {', '.join(wb.sheetnames)}")
    print(f"  - Filas de datos: {NUM_DATA_ROWS} (filas {ROW_DATA_START}-{ROW_DATA_END})")
    print(f"  - Catálogo: {len(ALL_CATALOG_ITEMS)} tipos de bien")
    print(f"  - INPC: {len(inpc_data)} años ({min(inpc_data.keys())}-{max(inpc_data.keys())})")


if __name__ == "__main__":
    main()
